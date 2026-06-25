import io
import logging
import os
import re
import time
import uuid
from typing import Optional

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    HTTPException,
    Query,
    Request,
    Response,
    UploadFile,
    status,
)
from fastapi import File as FileUpload
from open_webui.constants import ERROR_MESSAGES
from open_webui.env import (
    GLOSSARY_IMPORT_MAX_BYTES,
    GLOSSARY_IMPORT_TOKEN_TTL,
    SRC_LOG_LEVELS,
)
from open_webui.models.files import FileForm, Files
from open_webui.models.glossary import (
    Glossaries,
    GlossaryCopyForm,
    GlossaryForm,
    GlossaryResponse,
    GlossaryUserResponse,
    GovernanceMeta,
)
from open_webui.models.groups import Groups
from open_webui.models.models import Models
from open_webui.services.task_queue import InProcessQueue, TaskMessage
from open_webui.storage.provider import Storage
from open_webui.utils import glossary_file_parser as gfp
from open_webui.utils import glossary_llm_mapper as gllm
from open_webui.utils.access_control import (
    get_user_groups,
    has_access,
    has_permission_min_level,
)
from open_webui.utils.auth import get_verified_user
from open_webui.utils.license import require_feature
from pydantic import BaseModel, Field

log = logging.getLogger(__name__)
log.setLevel(SRC_LOG_LEVELS["MODELS"])

router = APIRouter(dependencies=[Depends(require_feature("glossary"))])


############################
# Search Engine Indexing Helper
############################


async def index_glossary_entries(app, glossary_id: str, entries: list[dict]):
    """용어집 entries 를 검색 엔진에 색인 (upsert).

    BackgroundTask 컨텍스트로 호출되므로 사용자 응답은 이미 끝난 상태. 인덱싱
    실패는 silent 하게 묻히면 DB(entries)와 search index 가 영구 불일치하므로
    ``Glossaries.set_index_state()`` 로 결과를 기록한다 (UI 알림 followup).

    빈 ``entry.id`` 는 skip — 동일 빈 키로 색인하면 마지막 doc 이 앞 doc 을
    덮어쓰는 silent data loss 가 발생한다.
    """
    if not entries:
        return

    valid_entries = [e for e in entries if e.get("id")]
    if len(valid_entries) < len(entries):
        # caller(bulk_replace/sync) 의 코드 버그 — 사용자 노출 에러 아니므로 warning
        log.warning(
            "[GLOSSARY_INDEX] skipping %d entries with empty id (glossary=%s)",
            len(entries) - len(valid_entries),
            glossary_id,
        )
    if not valid_entries:
        return

    log.info(
        "[GLOSSARY_INDEX] index_glossary_entries start: glossary=%s entries=%d",
        glossary_id,
        len(valid_entries),
    )

    try:
        from extension_modules.glossary import GlossaryEntryInput, GlossaryIndexService

        service = GlossaryIndexService(app)
        entry_inputs = [
            GlossaryEntryInput(
                id=entry["id"],
                glossary_id=glossary_id,
                term=entry.get("term", ""),
                synonyms=entry.get("synonyms", []),
                description=entry.get("description", ""),
                example=entry.get("example", ""),
                category=entry.get("category"),
            )
            for entry in valid_entries
        ]

        count = await service.index_entries(entry_inputs)
        total = len(entry_inputs)
        log.info(
            "[GLOSSARY_INDEX] index_glossary_entries done: indexed=%d/%d (glossary=%s)",
            count,
            total,
            glossary_id,
        )

        if count < total:
            # adapter 가 raise 하지 않고 일부 doc 만 succeeded=False 인 경우.
            # adapter 가 batch warning 으로 가시화하지만 caller side 도 meta 기록.
            Glossaries.set_index_state(
                glossary_id,
                {
                    "kind": "partial",
                    "indexed": count,
                    "total": total,
                    "ts": int(time.time()),
                },
            )
        else:
            Glossaries.set_index_state(glossary_id, None)

    except ValueError as e:
        # 검색 엔진 또는 임베딩이 미설정 — silent 가 정상 동작
        log.warning(
            "[GLOSSARY_INDEX] indexing skipped (config missing): glossary=%s entries=%d: %s",
            glossary_id,
            len(valid_entries),
            e,
        )
    except Exception as e:
        log.error(
            "[GLOSSARY_INDEX] FAILED to index %d entries (glossary=%s): %s",
            len(valid_entries),
            glossary_id,
            e,
            exc_info=True,
        )
        # 민감정보 보호: 전체 메시지는 log.error 에만, meta 에는 exception type +
        # 짧은 식별자만 남긴다 (Azure SDK 가 종종 endpoint URL/request id echo).
        Glossaries.set_index_state(
            glossary_id,
            {
                "kind": "error",
                "indexed": 0,
                "total": len(valid_entries),
                "error_type": type(e).__name__,
                "ts": int(time.time()),
            },
        )


async def delete_glossary_entries(app, entry_ids: list[str]):
    """
    특정 용어들을 검색 엔진에서 삭제.
    """
    if not entry_ids:
        return

    try:
        from extension_modules.glossary import GlossaryIndexService

        service = GlossaryIndexService(app)
        for entry_id in entry_ids:
            await service.delete_entry(entry_id)
        log.info(f"Deleted {len(entry_ids)} entries from search engine")

    except ValueError as e:
        log.warning(f"Glossary entry deletion skipped: {e}")
    except Exception as e:
        log.error(f"Failed to delete glossary entries: {e}")


async def delete_glossary_index(app, glossary_id: str):
    """
    용어집의 모든 용어를 검색 엔진에서 삭제.
    """
    try:
        from extension_modules.glossary import GlossaryIndexService

        service = GlossaryIndexService(app)
        count = await service.delete_by_glossary(glossary_id)
        log.info(f"Deleted {count} indexed entries for glossary {glossary_id}")

    except ValueError as e:
        log.warning(f"Glossary index deletion skipped: {e}")
    except Exception as e:
        log.error(f"Failed to delete glossary index: {e}")


async def sync_glossary_changes(
    app, glossary_id: str, old_entries: list[dict], new_entries: list[dict]
):
    """
    기존 entries와 새 entries를 비교하여 변경사항만 동기화.

    - 삭제된 entries: 검색 엔진에서 삭제
    - 추가/수정된 entries: 검색 엔진에 색인
    """
    old_ids = {e.get("id") for e in old_entries if e.get("id")}
    new_ids = {e.get("id") for e in new_entries if e.get("id")}

    # 삭제된 entries
    deleted_ids = old_ids - new_ids
    if deleted_ids:
        await delete_glossary_entries(app, list(deleted_ids))

    # 추가/수정된 entries (새 entries 전체를 upsert)
    if new_entries:
        await index_glossary_entries(app, glossary_id, new_entries)


############################
# getGlossaries
############################


def _strip_extract_job_result(glossary_dict: dict) -> dict:
    """목록 응답에서 무거운 extract_job.result 페이로드 제거 (status/progress 는 유지)."""
    meta = glossary_dict.get("meta")
    if not isinstance(meta, dict):
        return glossary_dict
    job = meta.get("extract_job")
    if not isinstance(job, dict) or "result" not in job:
        return glossary_dict
    new_meta = dict(meta)
    new_job = {k: v for k, v in job.items() if k != "result"}
    new_meta["extract_job"] = new_job
    glossary_dict["meta"] = new_meta
    return glossary_dict


def _build_governance_lookup(user_id: str) -> dict:
    """Request scope: 모든 Group dict + 사용자가 속한 group ids 1회 fetch.

    N+1 회피용. 라우터 함수 시작 시 1회 호출, 이후 hydrate 마다 재사용.
    """
    groups = {g.id: g.name for g in Groups.get_groups()}
    user_group_ids = {g.id for g in get_user_groups(user_id)}
    return {"groups": groups, "user_group_ids": user_group_ids}


def _hydrate_governance_meta(glossary, user, lookup: dict) -> GovernanceMeta:
    """access_control + 사용자 그룹 멤버십을 조합해 governance_meta 생성.

    정책 (Glossary Governance UX v2):
    - Group 단일축 노출. OU 직접 표시 안 함 (PR 156 정책).
    - 사용자 비멤버 그룹 이름은 노출 안 함 ("Shared" label, H1).
    - is_owner / can_write boolean 함께 derive (H3).
    """
    ac = glossary.access_control
    is_owner = glossary.user_id == user.id
    can_write = user.role == "admin" or is_owner or has_access(user.id, "write", ac)

    if ac is None:
        return GovernanceMeta(
            scope_kind="public",
            scope_label="Company-wide",
            is_owner=is_owner,
            can_write=can_write,
            is_my_group=False,
        )
    if not ac:  # 빈 dict → 비공개
        return GovernanceMeta(
            scope_kind="private",
            scope_label="Private",
            is_owner=is_owner,
            can_write=can_write,
            is_my_group=False,
        )

    read_groups = (ac.get("read") or {}).get("group_ids", []) or []
    write_groups = (ac.get("write") or {}).get("group_ids", []) or []
    referenced_groups = list(dict.fromkeys(read_groups + write_groups))

    user_group_ids = lookup["user_group_ids"]
    visible_groups = [gid for gid in referenced_groups if gid in user_group_ids]

    if visible_groups:
        steward_name = lookup["groups"].get(visible_groups[0])
        scope_label = steward_name or "Shared"
    else:
        steward_name = None
        scope_label = "Shared"

    is_my_group = bool(set(referenced_groups) & user_group_ids)

    return GovernanceMeta(
        scope_kind="group",
        scope_label=scope_label,
        steward_group_name=steward_name,
        is_owner=is_owner,
        can_write=can_write,
        is_my_group=is_my_group,
    )


@router.get("/", response_model=list[GlossaryUserResponse])
async def get_glossaries(request: Request, user=Depends(get_verified_user)):
    if user.role != "admin" and not has_permission_min_level(
        user.id,
        "workspace.glossaries",
        "read",
        request.app.state.config.USER_PERMISSIONS,
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
        )

    glossaries = []

    if user.role == "admin":
        glossaries = Glossaries.get_glossaries()
    else:
        glossaries = Glossaries.get_glossaries_by_user_id(user.id, "read")

    lookup = _build_governance_lookup(user.id)
    return [
        GlossaryUserResponse(
            **_strip_extract_job_result(glossary.model_dump()),
            governance_meta=_hydrate_governance_meta(glossary, user, lookup),
        )
        for glossary in glossaries
    ]


@router.get("/list", response_model=list[GlossaryUserResponse])
async def get_glossary_list(request: Request, user=Depends(get_verified_user)):
    # 편집 가능한 항목 목록 — write 권한자만 의미가 있음 (Schedules/KB 표준 패턴)
    if user.role != "admin" and not has_permission_min_level(
        user.id,
        "workspace.glossaries",
        "write",
        request.app.state.config.USER_PERMISSIONS,
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
        )

    glossaries = []

    if user.role == "admin":
        glossaries = Glossaries.get_glossaries()
    else:
        glossaries = Glossaries.get_glossaries_by_user_id(user.id, "write")

    lookup = _build_governance_lookup(user.id)
    return [
        GlossaryUserResponse(
            **_strip_extract_job_result(glossary.model_dump()),
            governance_meta=_hydrate_governance_meta(glossary, user, lookup),
        )
        for glossary in glossaries
    ]


############################
# CreateNewGlossary
############################


@router.post("/create", response_model=Optional[GlossaryResponse])
async def create_new_glossary(
    request: Request, form_data: GlossaryForm, user=Depends(get_verified_user)
):
    if user.role != "admin" and not has_permission_min_level(
        user.id,
        "workspace.glossaries",
        "write",
        request.app.state.config.USER_PERMISSIONS,
    ):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=ERROR_MESSAGES.UNAUTHORIZED,
        )

    if form_data.name and Glossaries.name_exists(form_data.name):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.NAME_TAKEN,
        )

    glossary = Glossaries.insert_new_glossary(user.id, form_data)

    if glossary:
        return glossary
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.DEFAULT("Failed to create glossary"),
        )


############################
# GetGlossaryCategories
############################


@router.get("/{id}/categories", response_model=list[str])
async def get_glossary_categories(
    id: str, request: Request, user=Depends(get_verified_user)
):
    """글로서리 카테고리 목록 반환.

    `meta.categories` 만 읽고 entries 컬럼은 건드리지 않아 5천 건 글로서리도
    즉시 응답한다. 권한 체크는 모델 메서드 안에서 수행 (None=글로서리 없음 또는 권한 없음).
    """
    if user.role != "admin" and not has_permission_min_level(
        user.id,
        "workspace.glossaries",
        "read",
        request.app.state.config.USER_PERMISSIONS,
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
        )

    cats = Glossaries.get_categories(
        id=id, user_id=user.id, is_admin=(user.role == "admin")
    )
    if cats is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=ERROR_MESSAGES.NOT_FOUND,
        )
    return cats


############################
# Category Rename / Delete (bulk on entries)
############################


class CategoryRenameForm(BaseModel):
    from_name: str = Field(..., min_length=1)
    to_name: str = Field(..., min_length=1)


@router.post("/{id}/categories/rename")
async def rename_glossary_category(
    id: str,
    request: Request,
    form_data: CategoryRenameForm,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    """카테고리 이름을 일괄 변경. extraction_sources 메타도 같이 이동."""
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    from_name = form_data.from_name.strip()
    to_name = form_data.to_name.strip()
    if not from_name or not to_name:
        raise HTTPException(status_code=400, detail="카테고리 이름은 비울 수 없습니다.")
    if from_name == to_name:
        raise HTTPException(
            status_code=400, detail="동일한 카테고리 이름으로 변경할 수 없습니다."
        )

    # 충돌 검증: to_name 이 다른 entry 에 이미 사용 중이면 거절
    entries = (glossary.data or {}).get("entries", []) or []
    if any(e.get("category") == to_name for e in entries):
        raise HTTPException(
            status_code=409, detail="이미 존재하는 카테고리 이름입니다."
        )

    changed = Glossaries.rename_category(id, from_name, to_name)
    if changed is None:
        raise HTTPException(status_code=500, detail="카테고리 이름 변경 실패")

    if changed:
        background_tasks.add_task(index_glossary_entries, request.app, id, changed)

    return {"success": True, "updated_count": len(changed)}


@router.delete("/{id}/categories/uncategorized")
async def delete_uncategorized_glossary_entries(
    id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    """카테고리가 없는(uncategorized) 모든 entries 를 일괄 삭제."""
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    removed = Glossaries.delete_category(id, None)
    if removed is None:
        raise HTTPException(status_code=500, detail="카테고리 삭제 실패")

    if removed:
        removed_ids = [e.get("id") for e in removed if e.get("id")]
        if removed_ids:
            background_tasks.add_task(delete_glossary_entries, request.app, removed_ids)

    return {"success": True, "deleted_count": len(removed)}


@router.delete("/{id}/categories/{name}")
async def delete_glossary_category(
    id: str,
    name: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    """카테고리와 그 안의 모든 entries 를 삭제. extraction_sources 메타도 정리.

    Deprecated: 카테고리명에 `/` 등 URL-unsafe 문자가 포함되면 path 라우팅이
    실패한다. 신규 호출은 ``POST /{id}/categories/delete`` 를 사용한다.
    """
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    name = (name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="카테고리 이름이 필요합니다.")

    removed = Glossaries.delete_category(id, name)
    if removed is None:
        raise HTTPException(status_code=500, detail="카테고리 삭제 실패")

    if removed:
        removed_ids = [e.get("id") for e in removed if e.get("id")]
        if removed_ids:
            background_tasks.add_task(delete_glossary_entries, request.app, removed_ids)

    return {"success": True, "deleted_count": len(removed)}


class CategoryDeleteForm(BaseModel):
    name: str = Field(..., min_length=1)
    keep_entries: bool = False


@router.post("/{id}/categories/delete")
async def delete_glossary_category_v2(
    id: str,
    request: Request,
    form_data: CategoryDeleteForm,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    """카테고리 삭제 (body 기반).

    Path parameter 가 아닌 body 로 이름을 받아 ``재무/결제`` 처럼 슬래시가 포함된
    카테고리명도 안전하게 처리한다.

    - ``keep_entries=False`` (기본): 해당 카테고리 + 안의 entries 모두 삭제.
    - ``keep_entries=True``: entries 는 미분류(uncategorized) 로 이동, 카테고리만 제거.
    """
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    name = (form_data.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="카테고리 이름이 필요합니다.")

    changed = Glossaries.delete_category(id, name, keep_entries=form_data.keep_entries)
    if changed is None:
        raise HTTPException(status_code=500, detail="카테고리 삭제 실패")

    if form_data.keep_entries:
        # entries 가 uncategorized 로 이동 — 카테고리 필드만 바뀌었으므로 reindex
        if changed:
            background_tasks.add_task(index_glossary_entries, request.app, id, changed)
        return {"success": True, "updated_count": len(changed), "kept_entries": True}

    if changed:
        removed_ids = [e.get("id") for e in changed if e.get("id")]
        if removed_ids:
            background_tasks.add_task(delete_glossary_entries, request.app, removed_ids)
    return {"success": True, "deleted_count": len(changed), "kept_entries": False}


############################
# Single Entry CRUD
############################


class GlossaryEntryCreateForm(BaseModel):
    term: str
    synonyms: Optional[list[str]] = None
    description: Optional[str] = ""
    example: Optional[str] = ""
    category: Optional[str] = None


class GlossaryEntryUpdateForm(BaseModel):
    term: Optional[str] = None
    synonyms: Optional[list[str]] = None
    description: Optional[str] = None
    example: Optional[str] = None
    category: Optional[str] = None


@router.post("/{id}/entries")
async def create_glossary_entry(
    id: str,
    request: Request,
    form_data: GlossaryEntryCreateForm,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    if not form_data.term.strip():
        raise HTTPException(status_code=400, detail="term is required")

    new_entry = Glossaries.add_entry(id, form_data.model_dump(), created_via="manual")
    if not new_entry:
        raise HTTPException(status_code=500, detail="Failed to add entry")

    background_tasks.add_task(index_glossary_entries, request.app, id, [new_entry])
    return new_entry


@router.put("/{id}/entries/{entry_id}")
async def update_glossary_entry(
    id: str,
    entry_id: str,
    request: Request,
    form_data: GlossaryEntryUpdateForm,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    patch = form_data.model_dump(exclude_none=True)
    updated = Glossaries.update_entry(id, entry_id, patch)
    if not updated:
        raise HTTPException(status_code=404, detail="Entry not found")

    background_tasks.add_task(index_glossary_entries, request.app, id, [updated])
    return updated


@router.delete("/{id}/entries/{entry_id}")
async def delete_glossary_entry(
    id: str,
    entry_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    removed = Glossaries.delete_entry(id, entry_id)
    if not removed:
        raise HTTPException(status_code=404, detail="Entry not found")

    background_tasks.add_task(delete_glossary_entries, request.app, [entry_id])
    return {"success": True, "id": entry_id}


class BulkDeleteEntriesForm(BaseModel):
    entry_ids: list[str] = Field(..., min_length=1, max_length=10000)


@router.post("/{id}/entries/bulk-delete")
async def bulk_delete_glossary_entries(
    id: str,
    request: Request,
    form_data: BulkDeleteEntriesForm,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    """체크박스 다중 선택 삭제. 단일 트랜잭션 + 일괄 검색 인덱스 제거."""
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    removed = Glossaries.delete_entries(id, form_data.entry_ids)
    if removed is None:
        raise HTTPException(status_code=500, detail="삭제 실패")

    if removed:
        removed_ids = [e.get("id") for e in removed if e.get("id")]
        if removed_ids:
            background_tasks.add_task(delete_glossary_entries, request.app, removed_ids)

    return {"success": True, "deleted_count": len(removed)}


############################
# Bulk Import (XLSX / CSV / MD)
############################

_IMPORT_TOKEN_KIND = "glossary_import"


def _check_glossary_feature_write(
    request: Request, user, detail: Optional[str] = None
) -> None:
    """Feature gate (workspace.glossaries write) → 401.

    Resource gate(403) 와 분리. admin 은 우회.
    """
    if user.role == "admin":
        return
    if not has_permission_min_level(
        user.id,
        "workspace.glossaries",
        "write",
        request.app.state.config.USER_PERMISSIONS,
    ):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=detail or ERROR_MESSAGES.UNAUTHORIZED,
        )


def _import_token_meta(user_id: str, glossary_id: str, file_size: int) -> dict:
    return {
        "kind": _IMPORT_TOKEN_KIND,
        "user_id": user_id,
        "glossary_id": glossary_id,
        "expires_at": int(time.time()) + GLOSSARY_IMPORT_TOKEN_TTL,
        "file_size": file_size,
    }


def _validate_import_token(token: str, user_id: str, glossary_id: str):
    """토큰 검증. file 반환 또는 HTTPException raise."""
    file = Files.get_file_by_id(token)
    if file is None:
        raise HTTPException(status_code=404, detail="Upload token not found")
    meta = file.meta or {}
    if meta.get("kind") != _IMPORT_TOKEN_KIND:
        raise HTTPException(status_code=404, detail="Upload token not found")
    if meta.get("user_id") != user_id:
        raise HTTPException(status_code=403, detail=ERROR_MESSAGES.ACCESS_PROHIBITED)
    if meta.get("glossary_id") != glossary_id:
        raise HTTPException(status_code=403, detail=ERROR_MESSAGES.ACCESS_PROHIBITED)
    if meta.get("expires_at", 0) < int(time.time()):
        # 410 Gone — 토큰 만료
        raise HTTPException(status_code=410, detail="Upload token expired")
    return file


def _load_file_bytes(file_record) -> bytes:
    """Files row 의 path 로 디스크/스토리지에서 bytes 로드."""
    path = file_record.path
    if not path:
        raise HTTPException(status_code=500, detail="저장된 파일 경로 없음")
    try:
        local_path = Storage.get_file(path)
        with open(local_path, "rb") as f:
            return f.read()
    except FileNotFoundError:
        raise HTTPException(status_code=410, detail="Upload token expired")
    except Exception as e:
        log.exception("import file load 실패: %s", e)
        raise HTTPException(status_code=500, detail="파일 로드 실패")


def _apply_llm_rule_to_rows(rule: dict, rows: list[dict]) -> tuple[list[dict], dict]:
    """LLM 변환 룰을 row 별로 적용 + 파일내 중복 term first-wins dedupe."""
    entries: list[dict] = []
    seen: set[str] = set()
    skipped = 0
    for row in rows:
        entry = gllm.apply_rule(rule, row)
        term = entry.get("term", "")
        if not term:
            skipped += 1
            continue
        key = term.lower()
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        entries.append(entry)
    return entries, {
        "total": len(rows),
        "parsed": len(entries),
        "skipped": skipped,
    }


def _delete_import_token(file_record) -> None:
    """토큰(=Files row) 폐기. 스토리지 + DB row 모두."""
    try:
        if file_record.path:
            Storage.delete_file(file_record.path)
    except Exception as e:
        log.warning("import token storage delete 실패: %s", e)
    try:
        Files.delete_file_by_id(file_record.id)
    except Exception as e:
        log.warning("import token DB delete 실패: %s", e)


# ============================================================
# Import 표준 템플릿 다운로드 (T9)
# ============================================================

_IMPORT_TEMPLATE_BYTES: Optional[bytes] = None
_IMPORT_TEMPLATE_ETAG = "glossary-import-template-v2"


def _build_import_template_xlsx() -> bytes:
    """5컬럼 + 예시 3행 표준 import 템플릿 .xlsx 바이트 생성. 모듈 캐시 대상."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "glossary"

    headers = ["term", "synonyms", "description", "example", "category"]
    # 예시 — 도메인 무관한 일반 비즈니스 약어 (실제 고객사 데이터 노출 회피).
    examples = [
        [
            "KPI",
            "Key Performance Indicator, 핵심성과지표",
            "비즈니스 목표 달성도를 측정하는 핵심 지표. 부서/팀/개인 단위로 설정한다.",
            "분기 KPI 검토 회의에서 매출 성장률을 보고한다.",
            "경영지표",
        ],
        [
            "API",
            "Application Programming Interface, 응용 프로그래밍 인터페이스",
            "서로 다른 소프트웨어 간 통신을 위한 정의된 규격.",
            "결제 시스템은 REST API 를 통해 외부 서비스와 연동된다.",
            "기술용어",
        ],
        [
            "SLA",
            "Service Level Agreement, 서비스 수준 협약",
            "서비스 제공자와 고객 간 서비스 품질 기준을 문서화한 합의.",
            "클라우드 SLA 는 가용성 99.9% 를 보장한다.",
            "계약",
        ],
    ]

    ws.append(headers)
    for row in examples:
        ws.append(row)

    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {"A": 15, "B": 35, "C": 60, "D": 50, "E": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # description / example 셀에 자동 줄바꿈 (긴 텍스트)
    for row_idx in range(2, 2 + len(examples)):
        for col_idx in (3, 4):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/import-template.xlsx")
async def download_glossary_import_template(
    request: Request,
    user=Depends(get_verified_user),
):
    """표준 import 템플릿 (.xlsx) 다운로드.

    - 5컬럼 + 예시 3행 (KPI / API / SLA — 일반 비즈니스 약어)
    - 모듈 캐시 + ETag (정적 콘텐츠).
    - Feature gate (workspace.glossaries read) — 다른 import endpoint 와 일관성.
    - glossary id 무관 (resource gate 불필요).
    """
    if user.role != "admin" and not has_permission_min_level(
        user.id,
        "workspace.glossaries",
        "read",
        request.app.state.config.USER_PERMISSIONS,
    ):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=ERROR_MESSAGES.UNAUTHORIZED,
        )

    global _IMPORT_TEMPLATE_BYTES
    if _IMPORT_TEMPLATE_BYTES is None:
        _IMPORT_TEMPLATE_BYTES = _build_import_template_xlsx()

    # If-None-Match → 304
    if request.headers.get("if-none-match") == _IMPORT_TEMPLATE_ETAG:
        return Response(status_code=304)

    return Response(
        content=_IMPORT_TEMPLATE_BYTES,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                'attachment; filename="glossary-import-template.xlsx"'
            ),
            "Cache-Control": "public, max-age=86400",
            "ETag": _IMPORT_TEMPLATE_ETAG,
        },
    )


@router.post("/{id}/import/preview")
async def preview_glossary_import(
    id: str,
    request: Request,
    user=Depends(get_verified_user),
    file: UploadFile = FileUpload(...),
):
    """업로드 파일을 파싱해 미리보기 + upload_token 발급.

    - Feature gate (workspace.glossaries write, 401) + Resource gate (403) 둘 다.
    - 파일 크기 cap (GLOSSARY_IMPORT_MAX_BYTES) 초과 시 413.
    - 파서 실패 (잘못된 헤더/포맷) 시 400.
    """
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_feature_write(request, user)
    _check_glossary_write_access(glossary, user)

    file_bytes = await file.read()
    if len(file_bytes) > GLOSSARY_IMPORT_MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"파일이 너무 큽니다 (한도: {GLOSSARY_IMPORT_MAX_BYTES} bytes).",
        )
    if not file_bytes:
        raise HTTPException(status_code=400, detail="파일이 비어 있습니다.")

    filename = file.filename or "glossary_import"
    try:
        parsed = gfp.parse_file(filename, file_bytes)
    except gfp.GlossaryParseError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    # 헤더 자동 매핑
    header_mapping = gfp.infer_header_mapping(parsed["headers"])

    # Sample preview (자동 매핑 + auto_classify OFF 기준)
    sample_parsed = gfp.ParsedFile(
        format=parsed["format"],
        headers=parsed["headers"],
        rows=parsed["rows"][:20],
        encoding=parsed["encoding"],
        md_pattern=parsed["md_pattern"],
    )
    sample_entries, _ = gfp.normalize_entries(
        sample_parsed, header_mapping, auto_classify=False
    )

    # 전체 통계 — 기존 entries 와 비교해 added/updated/skipped 계산
    full_entries, parse_stats = gfp.normalize_entries(
        parsed, header_mapping, auto_classify=False
    )
    existing_terms = {
        (e.get("term") or "").lower()
        for e in (glossary.data or {}).get("entries", [])
        if e.get("term")
    }
    added = sum(1 for e in full_entries if e["term"].lower() not in existing_terms)
    updated = len(full_entries) - added

    # Files 모델에 저장 — file_id = upload_token
    token = str(uuid.uuid4())
    # Path traversal 방어: 사용자 제공 filename (multipart 헤더) 을 sanitize.
    # basename 으로 디렉토리 분리 제거 + 영숫자/._- 외 문자 치환 + 길이 cap.
    safe_filename = (
        re.sub(r"[^A-Za-z0-9._\-]", "_", os.path.basename(filename))[:120] or "import"
    )
    storage_filename = f"glossary_import_{token}_{safe_filename}"
    try:
        _, storage_path = Storage.upload_file(io.BytesIO(file_bytes), storage_filename)
    except Exception as e:
        log.exception("import 파일 저장 실패: %s", e)
        raise HTTPException(status_code=500, detail="파일 저장 실패")

    file_record = Files.insert_new_file(
        user.id,
        FileForm(
            id=token,
            filename=filename,
            path=storage_path,
            meta=_import_token_meta(user.id, id, len(file_bytes)),
        ),
    )
    if file_record is None:
        try:
            Storage.delete_file(storage_path)
        except Exception:
            pass
        raise HTTPException(status_code=500, detail="토큰 생성 실패")

    return {
        "upload_token": token,
        "format": parsed["format"],
        "md_pattern": parsed["md_pattern"],
        "encoding": parsed["encoding"],
        "headers": parsed["headers"],
        "header_mapping": header_mapping,
        "sample_entries": sample_entries,
        "stats": {
            "total_rows": parse_stats["total"],
            "parsed": parse_stats["parsed"],
            "skipped_rows": parse_stats["skipped"],
            "added": added,
            "updated": updated,
        },
        "base_updated_at": glossary.updated_at,
    }


class CommitGlossaryImportForm(BaseModel):
    upload_token: str = Field(..., min_length=1)
    mapping: dict[str, str] = Field(default_factory=dict)
    auto_classify: bool = False
    base_updated_at: int = Field(..., ge=0)
    # LLM-aided rule (선택). 있으면 mapping/auto_classify 대신 deterministic apply.
    llm_rule: Optional[dict] = None


class LLMSuggestForm(BaseModel):
    upload_token: str = Field(..., min_length=1)
    model_id: Optional[str] = None
    # Hallucination 방어 옵션: description 의 LLM 합성문 (template/constant) 금지.
    # ON 시 description spec 은 column/concat/skip 으로만 결과 생성됨.
    conservative_description: bool = False


@router.post("/{id}/import/llm-suggest")
async def llm_suggest_glossary_import(
    id: str,
    request: Request,
    form_data: LLMSuggestForm,
    user=Depends(get_verified_user),
):
    """업로드된 raw 자료의 변환 룰을 LLM 으로 추론.

    - 권한 (feature 401 + resource 403)
    - 토큰 검증
    - 파일 재로드 + sample 5~8 row 추출 → LLM 호출 → 룰 JSON 반환
    - sample 에 룰을 미리 적용한 결과도 함께 반환 (preview)
    - LLM 호출 실패는 502, 룰 검증 실패는 400
    """
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_feature_write(request, user)
    _check_glossary_write_access(glossary, user)

    file_record = _validate_import_token(form_data.upload_token, user.id, id)
    file_bytes = _load_file_bytes(file_record)
    try:
        parsed = gfp.parse_file(file_record.filename or "import", file_bytes)
    except gfp.GlossaryParseError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    sample_rows = parsed["rows"][:8]
    try:
        rule = await gllm.suggest_rule(
            request.app,
            form_data.model_id or "",
            parsed["headers"],
            sample_rows,
            conservative_description=form_data.conservative_description,
        )
    except gllm.RuleValidationError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception as e:
        log.exception("LLM 룰 추론 실패: %s", e)
        raise HTTPException(status_code=502, detail=f"LLM 호출 실패: {e}") from e

    # sample 에 룰 적용 — 사용자 미리보기용 + segment 시각 표시
    sample_entries, _ = _apply_llm_rule_to_rows(rule, sample_rows)
    # Hallucination 방어 A — 각 sample entry 의 segment (literal=AI 작성, data=파일 값)
    sample_segments: list[dict] = []
    for row in sample_rows[: len(sample_entries)]:
        try:
            result = gllm.apply_rule_with_segments(rule, row)
            sample_segments.append(result.get("segments", {}))
        except Exception as e:
            log.warning("segment 분해 실패 (UI 시각 표시만 영향): %s", e)
            sample_segments.append({})
    return {
        "rule": rule,
        "sample_entries": sample_entries[:20],
        "sample_entries_segments": sample_segments[:20],
    }


@router.post("/{id}/import/commit")
async def commit_glossary_import(
    id: str,
    request: Request,
    form_data: CommitGlossaryImportForm,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    """미리보기에서 발급된 토큰으로 실제 등록.

    - 권한 재검증 (feature 401 + resource 403)
    - 토큰 검증 (404 / 403 cross-tenant / 410 expired)
    - extract_job 진행 중 → 409
    - optimistic version (updated_at 비교) → 412
    - merge → 인덱싱 (BackgroundTasks)
    - 토큰 폐기 (Storage + Files row)
    """
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_feature_write(request, user)
    _check_glossary_write_access(glossary, user)

    file_record = _validate_import_token(form_data.upload_token, user.id, id)

    # extract_job 충돌 — 사용자가 곧 다시 commit 시도할 가능성 있으니 토큰 유지
    extract_job = (glossary.meta or {}).get("extract_job") or {}
    if extract_job.get("status") in ("queued", "running"):
        raise HTTPException(
            status_code=409,
            detail="추출 작업이 진행 중입니다. 완료 또는 취소 후 다시 시도하세요.",
        )

    # Optimistic version 불일치 — 베이스 변경됨. 토큰 폐기 + 사용자 재시도 유도.
    if glossary.updated_at != form_data.base_updated_at:
        _delete_import_token(file_record)
        raise HTTPException(
            status_code=412,
            detail="다른 사용자가 용어집을 수정했습니다. 새로고침 후 다시 시도하세요.",
        )

    # 파일 재로드 + 전체 파싱
    file_bytes = _load_file_bytes(file_record)
    try:
        parsed = gfp.parse_file(file_record.filename or "import", file_bytes)
    except gfp.GlossaryParseError as e:
        _delete_import_token(file_record)
        raise HTTPException(status_code=400, detail=str(e)) from e

    if form_data.llm_rule:
        # LLM-aided 변환 룰 적용 경로 — mapping/auto_classify 무시
        try:
            rule = gllm.validate_rule(form_data.llm_rule)
        except gllm.RuleValidationError as e:
            _delete_import_token(file_record)
            raise HTTPException(
                status_code=400, detail=f"변환 룰 형식 오류: {e}"
            ) from e
        entries, parse_stats = _apply_llm_rule_to_rows(rule, parsed["rows"])
    else:
        # mapping: 사용자 수정본이 비어있으면 자동 추론으로 fallback
        mapping = form_data.mapping or gfp.infer_header_mapping(parsed["headers"])
        entries, parse_stats = gfp.normalize_entries(
            parsed, mapping, auto_classify=form_data.auto_classify
        )
    if not entries:
        _delete_import_token(file_record)
        raise HTTPException(
            status_code=400,
            detail="등록할 용어가 없습니다. 매핑 또는 파일을 확인하세요.",
        )

    # Hallucination 검수 워크플로우 (T14): 신규 entry 에 created_via 부여.
    # llm_rule 사용 시 'ai_rule', 단순 매핑이면 'file_import_no_ai'.
    new_created_via = "ai_rule" if form_data.llm_rule else "file_import_no_ai"

    # 신규/업데이트 분리 + entry id 부여
    existing = list((glossary.data or {}).get("entries", []))
    existing_by_term: dict[str, dict] = {
        (e.get("term") or "").lower(): e for e in existing if e.get("term")
    }
    now = int(time.time())
    added_count = 0
    updated_count = 0
    final_entries: list[dict] = list(existing)
    for new_entry in entries:
        key = new_entry["term"].lower()
        if key in existing_by_term:
            old = existing_by_term[key]
            old_id = old.get("id")
            merged = {
                **old,
                **new_entry,
                "id": old_id,
                # 기존 created_via 보존 — update 는 출처를 바꾸지 않음
                "created_via": old.get("created_via") or new_created_via,
                "updated_at": now,
            }
            # remaining 에서 old 를 swap
            for i, e in enumerate(final_entries):
                if e.get("id") == old_id:
                    final_entries[i] = merged
                    break
            updated_count += 1
        else:
            new_entry_with_id = {
                **new_entry,
                "id": str(uuid.uuid4()),
                "created_via": new_created_via,
                "created_at": now,
                "updated_at": now,
            }
            final_entries.append(new_entry_with_id)
            added_count += 1

    # data 갱신 (update_glossary_data_by_id 가 meta.categories 자동 재계산)
    updated_glossary = Glossaries.update_glossary_data_by_id(
        id, {"entries": final_entries}
    )
    if updated_glossary is None:
        raise HTTPException(status_code=500, detail="등록 실패")

    # 토큰 폐기
    _delete_import_token(file_record)

    # 검색 인덱싱 (BackgroundTasks). 변경된 entries 만 reindex.
    changed_for_index = [e for e in final_entries if e.get("updated_at") == now]
    if changed_for_index:
        background_tasks.add_task(
            index_glossary_entries, request.app, id, changed_for_index
        )

    return {
        "success": True,
        "added": added_count,
        "updated": updated_count,
        "skipped": parse_stats["skipped"],
        "total": len(entries),
    }


############################
# GetGlossaryEntriesPaginated
############################


@router.get("/{id}/entries")
async def get_glossary_entries_paginated(
    id: str,
    skip: int = 0,
    limit: int = 50,
    search: Optional[str] = None,
    sort: str = "name",
    category: Optional[str] = None,
    uncategorized: bool = False,
    categories: Optional[list[str]] = Query(default=None),
    include_uncategorized: bool = False,
    created_via: Optional[list[str]] = Query(default=None),
    user=Depends(get_verified_user),
):
    """용어집 entries 를 skip/limit 으로 페이징. 검색/정렬/카테고리 서버에서 처리.

    카테고리 필터 (우선순위):
    - ``categories`` (반복 query param, ex: ``?categories=A&categories=B``) — 멀티 셀렉트.
      ``include_uncategorized=true`` 와 함께 쓰면 카테고리 없는 entries 도 포함.
    - ``uncategorized=true`` — 카테고리 없는 것만.
    - ``category`` — 단일 카테고리 (legacy).
    """
    if limit <= 0 or limit > 500:
        limit = 50
    if skip < 0:
        skip = 0
    if sort not in ("name", "newest", "oldest"):
        sort = "name"

    glossary = Glossaries.get_glossary_by_id(id=id)
    if not glossary:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail=ERROR_MESSAGES.NOT_FOUND
        )
    if not (
        user.role == "admin"
        or glossary.user_id == user.id
        or has_access(user.id, "read", glossary.access_control)
    ):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail=ERROR_MESSAGES.UNAUTHORIZED
        )

    result = Glossaries.get_entries_paginated(
        id=id,
        skip=skip,
        limit=limit,
        search=search,
        sort=sort,
        category=category,
        uncategorized=uncategorized,
        categories=categories,
        include_uncategorized=include_uncategorized,
        created_via=created_via,
    )
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail=ERROR_MESSAGES.NOT_FOUND
        )
    return {
        "entries": result["entries"],
        "total": result["total"],
        "skip": skip,
        "limit": limit,
    }


############################
# GetGlossaryById
############################


@router.get("/{id}", response_model=Optional[GlossaryResponse])
async def get_glossary_by_id(id: str, user=Depends(get_verified_user)):
    glossary = Glossaries.get_glossary_by_id(id=id)

    if glossary:
        if (
            user.role == "admin"
            or glossary.user_id == user.id
            or has_access(user.id, "read", glossary.access_control)
        ):
            return GlossaryResponse(**glossary.model_dump())

    raise HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail=ERROR_MESSAGES.NOT_FOUND,
    )


############################
# UpdateGlossaryById
############################


@router.post("/{id}/update", response_model=Optional[GlossaryResponse])
async def update_glossary_by_id(
    id: str,
    form_data: GlossaryForm,
    request: Request,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    glossary = Glossaries.get_glossary_by_id(id=id)
    if not glossary:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.NOT_FOUND,
        )

    if (
        glossary.user_id != user.id
        and not has_access(user.id, "write", glossary.access_control)
        and user.role != "admin"
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
        )

    if form_data.name and Glossaries.name_exists(form_data.name, exclude_id=id):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.NAME_TAKEN,
        )

    # 기존 entries 저장 (검색 엔진 동기화용)
    old_entries = (glossary.data or {}).get("entries", [])
    new_entries = (form_data.data or {}).get("entries", []) if form_data.data else None

    glossary = Glossaries.update_glossary_by_id(id=id, form_data=form_data)
    if glossary:
        # entries가 변경된 경우 검색 엔진 동기화 (백그라운드)
        if new_entries is not None:
            background_tasks.add_task(
                sync_glossary_changes, request.app, id, old_entries, new_entries
            )

        return GlossaryResponse(**glossary.model_dump())
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.DEFAULT("Failed to update glossary"),
        )


############################
# SyncGlossaryToSearchEngine
############################


@router.post("/{id}/sync")
async def sync_glossary_to_search_engine(
    id: str,
    request: Request,
    user=Depends(get_verified_user),
):
    """용어집 용어를 검색 엔진에 동기화"""
    glossary = Glossaries.get_glossary_by_id(id=id)
    if not glossary:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.NOT_FOUND,
        )

    if (
        glossary.user_id != user.id
        and not has_access(user.id, "write", glossary.access_control)
        and user.role != "admin"
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
        )

    entries = (glossary.data or {}).get("entries", [])
    if not entries:
        return {"status": True, "indexed": 0, "message": "No entries to sync"}

    try:
        await index_glossary_entries(request.app, id, entries)
        return {
            "status": True,
            "indexed": len(entries),
            "message": f"Successfully synced {len(entries)} terms to search engine",
        }
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e),
        )
    except Exception as e:
        log.error(f"Failed to sync glossary: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to sync: {str(e)}",
        )


############################
# GetLinkedAgentsByGlossaryId
############################


@router.get("/{id}/linked-agents")
async def get_linked_agents_by_glossary_id(
    id: str, request: Request, user=Depends(get_verified_user)
):
    """Return agents (models) that have this glossary connected."""
    if user.role != "admin" and not has_permission_min_level(
        user.id,
        "workspace.glossaries",
        "read",
        request.app.state.config.USER_PERMISSIONS,
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
        )

    glossary = Glossaries.get_glossary_by_id(id=id)
    if not glossary:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=ERROR_MESSAGES.NOT_FOUND,
        )

    if (
        user.role != "admin"
        and glossary.user_id != user.id
        and not has_access(user.id, "read", glossary.access_control)
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
        )

    models = Models.get_all_models()
    linked = []
    for model in models:
        if not model.meta:
            continue
        meta = model.meta.model_dump() if hasattr(model.meta, "model_dump") else {}
        glossaries = meta.get("glossaries", []) or []
        if isinstance(glossaries, str):
            glossaries = [glossaries]
        for g in glossaries:
            if isinstance(g, dict) and g.get("id") == id:
                linked.append({"id": model.id, "name": model.name})
                break
            elif isinstance(g, str) and g == id:
                linked.append({"id": model.id, "name": model.name})
                break
    return linked


############################
# DeleteGlossaryById
############################


@router.delete("/{id}/delete", response_model=bool)
async def delete_glossary_by_id(
    id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    glossary = Glossaries.get_glossary_by_id(id=id)
    if not glossary:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.NOT_FOUND,
        )

    if (
        glossary.user_id != user.id
        and not has_access(user.id, "write", glossary.access_control)
        and user.role != "admin"
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
        )

    log.info(f"Deleting glossary: {id} (name: {glossary.name})")

    result = Glossaries.delete_glossary_by_id(id=id)

    if result:
        # 검색 엔진에서 용어 삭제 (백그라운드)
        background_tasks.add_task(delete_glossary_index, request.app, id)

    return result


############################
# CopyGlossary (governance fork)
############################


@router.post("/{id}/copy", response_model=Optional[GlossaryResponse])
async def copy_glossary(
    id: str,
    request: Request,
    form_data: GlossaryCopyForm,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    """글로서리 fork — 사용자가 멤버인 Group 또는 비공개로 복사.

    정책 (Glossary Governance UX v2):
    - 카테고리 write 권한 필수 (workspace.glossaries=write)
    - 소유자 / admin / write 공유자만 호출 가능 (read 공유자 차단)
    - target_group_id 는 사용자 멤버 그룹 중 하나여야 함 (admin 예외)
    - target_group_id 없음 = 비공개({}) 강제
    - entry id 모두 재발급 + meta.copied_from 기록
    - BackgroundTasks 로 검색 재인덱싱
    """
    if user.role != "admin" and not has_permission_min_level(
        user.id,
        "workspace.glossaries",
        "write",
        request.app.state.config.USER_PERMISSIONS,
    ):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=ERROR_MESSAGES.UNAUTHORIZED,
        )

    glossary = Glossaries.get_glossary_by_id(id=id)
    # 소유자 / admin / write 공유자만 (read 공유자 차단). 헬퍼가 404/403 처리.
    _check_glossary_write_access(glossary, user)

    # target_group_id 검증: 사용자 멤버여야 (admin 예외)
    if form_data.target_group_id and user.role != "admin":
        user_group_ids = {g.id for g in get_user_groups(user.id)}
        if form_data.target_group_id not in user_group_ids:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
            )

    new_glossary = Glossaries.copy_glossary(
        source_id=id,
        new_user_id=user.id,
        new_user_name=user.name,
        name=form_data.name,
        target_group_id=form_data.target_group_id,
    )
    if not new_glossary:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.DEFAULT("Failed to copy glossary"),
        )

    # 새 glossary_id 로 entries 검색 재인덱싱 (entry id 모두 재발급됐음)
    entries = (new_glossary.data or {}).get("entries", []) or []
    if entries:
        background_tasks.add_task(
            index_glossary_entries, request.app, new_glossary.id, entries
        )

    return new_glossary


############################
# Extract Values from DB
############################


class GlossaryCountValuesForm(BaseModel):
    dbsphere_id: str
    table_name: str
    column_name: str


@router.post("/{id}/count-values")
async def count_distinct_values(
    id: str,
    request: Request,
    form_data: GlossaryCountValuesForm,
    user=Depends(get_verified_user),
):
    """DB 컬럼의 DISTINCT 값 개수만 빠르게 반환한다."""
    import copy

    from extension_modules.dbsphere.dbsphere_state import DBConfig, DBType
    from extension_modules.dbsphere.sql_runners.databricks import DatabricksRunner
    from extension_modules.dbsphere.sql_runners.fabric import FabricRunner
    from extension_modules.dbsphere.sql_runners.mssql import MSSQLRunner
    from extension_modules.dbsphere.sql_runners.mysql import MySQLRunner
    from extension_modules.dbsphere.sql_runners.oracle import OracleRunner
    from extension_modules.dbsphere.sql_runners.postgres import PostgresRunner
    from extension_modules.dbsphere.sql_runners.snowflake import SnowflakeRunner
    from extension_modules.dbsphere.sql_runners.synapse import SynapseRunner
    from open_webui.models.dbsphere import DbSpheres
    from open_webui.routers.dbsphere import decrypt_connection_password

    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    dbsphere = DbSpheres.get_dbsphere_by_id(form_data.dbsphere_id)
    if not dbsphere or not dbsphere.data:
        raise HTTPException(status_code=404, detail="DbSphere not found")

    data = decrypt_connection_password(copy.deepcopy(dbsphere.data))
    db_config = DBConfig.from_dbsphere_data(data)
    db_type = db_config.get_db_type_enum()

    runner_map = {
        DBType.POSTGRES: PostgresRunner,
        DBType.MYSQL: MySQLRunner,
        DBType.MSSQL: MSSQLRunner,
        DBType.ORACLE: OracleRunner,
        DBType.SNOWFLAKE: SnowflakeRunner,
        DBType.DATABRICKS: DatabricksRunner,
        DBType.SYNAPSE: SynapseRunner,
        DBType.FABRIC: FabricRunner,
    }
    runner_cls = runner_map.get(db_type)
    if not runner_cls:
        raise HTTPException(status_code=400, detail=f"Unsupported DB type: {db_type}")
    runner = runner_cls(db_config)

    q = "`" if db_type in (DBType.MYSQL, DBType.DATABRICKS) else '"'
    col = f"{q}{form_data.column_name}{q}"
    tbl = f"{q}{form_data.table_name}{q}"
    sql = f"SELECT COUNT(DISTINCT {col}) FROM {tbl} WHERE {col} IS NOT NULL"

    try:
        df = await runner.run_sql(sql)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SQL error: {e}")

    count = int(df.iloc[0, 0]) if len(df) > 0 else 0
    return {"distinct_count": count}


class GlossaryExtractValuesForm(BaseModel):
    dbsphere_id: str
    table_name: str
    column_name: str
    synonym_column: Optional[str] = None
    description_column: Optional[str] = None
    context_columns: Optional[list[str]] = None
    model_id: Optional[str] = None
    generate_enrichment: bool = True
    llm_fields: Optional[list[str]] = None  # ["synonyms", "description", "example"]
    batch_size: int = 10
    llm_concurrency: int = 8  # 동시 LLM 호출 수 (TPM 에 맞춰 조정)
    category: str = Field(..., min_length=1)
    custom_instructions: Optional[str] = None  # User instructions for LLM enrichment


def _check_glossary_write_access(glossary, user) -> None:
    if not glossary:
        raise HTTPException(status_code=404, detail=ERROR_MESSAGES.NOT_FOUND)
    if (
        glossary.user_id != user.id
        and user.role != "admin"
        and not has_access(user.id, "write", glossary.access_control)
    ):
        raise HTTPException(status_code=403, detail=ERROR_MESSAGES.ACCESS_PROHIBITED)


@router.post("/{id}/extract-values")
async def start_extract_values_job(
    id: str,
    request: Request,
    form_data: GlossaryExtractValuesForm,
    user=Depends(get_verified_user),
):
    """DB 테이블 컬럼의 DISTINCT 값을 백그라운드로 추출 → LLM enrichment → preview.

    즉시 `{job_id, status}` 를 반환하고, 컨슈머가 작업을 처리하면서
    `glossary.meta.extract_job` 을 갱신한다. 완료 시 Socket.IO 알림 전송.
    """
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    existing_job = (glossary.meta or {}).get("extract_job") if glossary else None
    if existing_job and existing_job.get("status") in ("queued", "running"):
        raise HTTPException(
            status_code=409,
            detail="이미 진행 중인 추출 작업이 있습니다.",
        )

    # LLM 모델 설정은 **HTTP 요청 컨텍스트**에서 미리 resolve 해서 payload 에
    # 넣는다. 백그라운드 consumer 컨텍스트에서는 app.state.MODELS 가 아직
    # populate 되지 않았을 수 있어 엉뚱한 provider 로 매칭되는 사고 방지.
    pre_resolved_model_config = None
    if form_data.generate_enrichment and form_data.model_id:
        from extension_modules.utils.llm import (
            get_model_config_from_app,
            model_config_has_credentials,
        )

        try:
            pre_resolved_model_config = get_model_config_from_app(
                request.app, form_data.model_id
            )
        except Exception as e:
            log.warning(f"[glossary extract] pre-resolve model config failed: {e}")
        if pre_resolved_model_config is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"모델 설정을 찾을 수 없습니다: {form_data.model_id}. "
                    "관리자 페이지에서 모델이 활성화되어 있는지 확인해 주세요."
                ),
            )
        if not model_config_has_credentials(pre_resolved_model_config):
            raise HTTPException(
                status_code=400,
                detail=(
                    f"선택한 모델에 자격증명이 없습니다: {form_data.model_id}. "
                    "(openai/azure 는 api_key, vertex 는 service_account_key 또는 "
                    "use_global_gcp_key 필요)"
                ),
            )

    job_id = str(uuid.uuid4())
    now = int(time.time())
    new_job = {
        "id": job_id,
        "status": "queued",
        "phase": "queued",
        "progress": {"current": 0, "total": 0},
        "params": form_data.model_dump(),
        "started_at": None,
        "completed_at": None,
        "error": None,
        "result": None,
        "queued_at": now,
    }
    Glossaries.set_extract_job(id, new_job)

    payload = {
        "glossary_id": id,
        "user_id": user.id,
        "params": form_data.model_dump(),
        "pre_resolved_model_config": pre_resolved_model_config,
    }

    task_queue = getattr(request.app.state, "task_queue", None)
    if task_queue is None or isinstance(task_queue, InProcessQueue):
        # Redis 가 없는 환경: 같은 프로세스에서 즉시 실행 (개발 편의)
        import asyncio as _asyncio

        from extension_modules.glossary.extract_worker import (
            process_glossary_extract_task,
        )

        _asyncio.create_task(
            process_glossary_extract_task(request.app, job_id, payload)
        )
    else:
        await task_queue.publish(
            TaskMessage(
                task_type="glossary_extract_values",
                payload=payload,
                task_id=job_id,
            )
        )

    return {"job_id": job_id, "status": "queued"}


@router.get("/{id}/extract-job")
async def get_extract_job(id: str, user=Depends(get_verified_user)):
    """현재 잡 상태(result 포함)를 반환. 잡이 없으면 null."""
    glossary = Glossaries.get_glossary_by_id(id=id)
    if not glossary:
        raise HTTPException(status_code=404, detail=ERROR_MESSAGES.NOT_FOUND)
    if (
        glossary.user_id != user.id
        and user.role != "admin"
        and not has_access(user.id, "read", glossary.access_control)
    ):
        raise HTTPException(status_code=403, detail=ERROR_MESSAGES.ACCESS_PROHIBITED)
    return (glossary.meta or {}).get("extract_job")


@router.get("/extract-jobs/active")
async def list_active_extract_jobs(user=Depends(get_verified_user)):
    """현재 사용자가 접근 가능한 용어집 중 활성 추출 잡 목록 (요약)."""
    return Glossaries.get_active_extract_jobs_by_user(user.id)


def _merge_entries_overwrite(
    existing: list[dict],
    to_apply: list[dict],
) -> list[dict]:
    """term(lowercase) 기준으로 덮어쓰기 병합. 신규는 append.

    extract-job 경로 (accept / accept-all) 에서만 호출 — incoming entry 의
    ``created_via`` 가 비어있으면 'extract_db' 로 채운다.
    """
    by_term: dict[str, dict] = {}
    order: list[str] = []
    for e in existing:
        key = (e.get("term") or "").lower()
        by_term[key] = e
        order.append(key)
    now = int(time.time())
    for new in to_apply:
        key = (new.get("term") or "").lower()
        if not key:
            continue
        merged = dict(new)
        if key in by_term:
            merged["created_at"] = by_term[key].get("created_at") or now
            # update — 기존 created_via 보존
            merged["created_via"] = (
                by_term[key].get("created_via")
                or merged.get("created_via")
                or "extract_db"
            )
        else:
            # 신규 — incoming 명시 없으면 extract_db (이 함수의 유일한 caller 컨텍스트)
            if not merged.get("created_via"):
                merged["created_via"] = "extract_db"
        merged["updated_at"] = now
        if key in by_term:
            by_term[key] = merged
        else:
            by_term[key] = merged
            order.append(key)
    return [by_term[k] for k in order]


def _pop_pending_entry(job: dict, entry_id: str) -> Optional[dict]:
    result = job.get("result") or {}
    entries = list(result.get("entries") or [])
    found = None
    remaining = []
    for e in entries:
        if e.get("id") == entry_id and found is None:
            found = e
        else:
            remaining.append(e)
    if found is None:
        return None
    result["entries"] = remaining
    job["result"] = result
    return found


def _maybe_clear_empty_job(glossary_id: str, job: dict) -> None:
    """남은 pending entry 가 없으면 잡 전체 제거."""
    result = job.get("result") or {}
    entries = result.get("entries") or []
    if not entries:
        Glossaries.set_extract_job(glossary_id, None)
    else:
        Glossaries.set_extract_job(glossary_id, job)


@router.post("/{id}/extract-job/entries/{entry_id}/accept")
async def accept_extract_entry(
    id: str,
    entry_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    job = (glossary.meta or {}).get("extract_job") if glossary else None
    if not job or job.get("status") != "succeeded":
        raise HTTPException(status_code=400, detail="적용 가능한 잡이 없습니다.")

    job = dict(job)
    entry = _pop_pending_entry(job, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="해당 미리보기 항목이 없습니다.")

    existing_entries = list((glossary.data or {}).get("entries") or [])
    merged = _merge_entries_overwrite(existing_entries, [entry])
    new_data = dict(glossary.data or {})
    new_data["entries"] = merged
    Glossaries.update_glossary_data_by_id(id, new_data)
    _maybe_clear_empty_job(id, job)

    background_tasks.add_task(
        sync_glossary_changes, request.app, id, existing_entries, merged
    )

    return Glossaries.get_glossary_by_id(id=id)


@router.post("/{id}/extract-job/entries/{entry_id}/reject")
async def reject_extract_entry(id: str, entry_id: str, user=Depends(get_verified_user)):
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    job = (glossary.meta or {}).get("extract_job") if glossary else None
    if not job or job.get("status") != "succeeded":
        raise HTTPException(status_code=400, detail="처리 가능한 잡이 없습니다.")

    job = dict(job)
    entry = _pop_pending_entry(job, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="해당 미리보기 항목이 없습니다.")
    _maybe_clear_empty_job(id, job)
    return Glossaries.get_glossary_by_id(id=id)


class GlossaryExtractEntryUpdateForm(BaseModel):
    term: Optional[str] = None
    synonyms: Optional[list[str]] = None
    description: Optional[str] = None
    example: Optional[str] = None
    category: Optional[str] = None


@router.put("/{id}/extract-job/entries/{entry_id}")
async def update_extract_entry(
    id: str,
    entry_id: str,
    form_data: GlossaryExtractEntryUpdateForm,
    user=Depends(get_verified_user),
):
    """미리보기 단계의 pending entry 를 편집한다 (data.entries 에는 아직 안 들어감)."""
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    job = (glossary.meta or {}).get("extract_job") if glossary else None
    if not job or job.get("status") != "succeeded":
        raise HTTPException(status_code=400, detail="처리 가능한 잡이 없습니다.")

    job = dict(job)
    result = dict(job.get("result") or {})
    entries = list(result.get("entries") or [])
    patch = form_data.model_dump(exclude_none=True)
    found = False
    for i, e in enumerate(entries):
        if e.get("id") == entry_id:
            updated = dict(e)
            updated.update(patch)
            updated["updated_at"] = int(time.time())
            entries[i] = updated
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail="해당 미리보기 항목이 없습니다.")
    result["entries"] = entries
    job["result"] = result
    Glossaries.set_extract_job(id, job)
    return Glossaries.get_glossary_by_id(id=id)


@router.post("/{id}/extract-job/accept-all")
async def accept_all_extract_entries(
    id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user=Depends(get_verified_user),
):
    """모든 pending entries 를 일괄 병합(덮어쓰기) 후 잡 종료."""
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    job = (glossary.meta or {}).get("extract_job") if glossary else None
    if not job or job.get("status") != "succeeded":
        raise HTTPException(status_code=400, detail="적용 가능한 잡이 없습니다.")

    pending_entries = list((job.get("result") or {}).get("entries") or [])
    if not pending_entries:
        Glossaries.set_extract_job(id, None)
        return Glossaries.get_glossary_by_id(id=id)

    existing_entries = list((glossary.data or {}).get("entries") or [])
    merged = _merge_entries_overwrite(existing_entries, pending_entries)
    new_data = dict(glossary.data or {})
    new_data["entries"] = merged
    Glossaries.update_glossary_data_by_id(id, new_data)
    Glossaries.set_extract_job(id, None)

    background_tasks.add_task(
        sync_glossary_changes, request.app, id, existing_entries, merged
    )

    return Glossaries.get_glossary_by_id(id=id)


@router.post("/{id}/extract-job/discard")
async def discard_extract_job(id: str, user=Depends(get_verified_user)):
    """미리보기 결과를 통째로 버리고 잡을 종료한다."""
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)
    Glossaries.set_extract_job(id, None)
    return Glossaries.get_glossary_by_id(id=id)


@router.post("/{id}/extract-job/cancel")
async def cancel_extract_job(id: str, user=Depends(get_verified_user)):
    """진행 중인 잡 취소 신호. 워커가 다음 배치 시작 전 감지해서 중단."""
    glossary = Glossaries.get_glossary_by_id(id=id)
    _check_glossary_write_access(glossary, user)

    job = (glossary.meta or {}).get("extract_job") if glossary else None
    if not job or job.get("status") not in ("queued", "running"):
        raise HTTPException(status_code=400, detail="진행 중인 잡이 없습니다.")
    Glossaries.patch_extract_job(
        id, {"status": "canceled", "completed_at": int(time.time())}
    )
    return Glossaries.get_glossary_by_id(id=id)
