"""Glossary 일괄 import 용 파일 파서.

지원 포맷: XLSX / CSV / Markdown.

- XLSX: openpyxl `read_only=True, data_only=True` 스트리밍. formula/macro 무시.
- CSV: stdlib csv + 인코딩 fallback chain (utf-8-sig → utf-8 → cp949 → charset-normalizer).
- MD: markdown-it-py token AST 기반 (table) + regex 기반 (sectioned, h2/dl).

모든 파서는 ``ParsedFile`` 을 반환 — preview 응답과 commit 단계에서 동일하게 사용된다.

DoS 가드:
- ``MAX_ROWS`` (50,000) — XLSX/CSV row count cap.
- ``MAX_SHEET_CELLS`` (5,000,000) — XLSX sheet 셀 수 cap.
- 파일 크기 cap 은 router 단에서 ``GLOSSARY_IMPORT_MAX_BYTES`` 로 별도 강제.

본 모듈은 ``backend/open_webui/routers/glossary.py`` 의 import preview/commit endpoint
에서 호출된다.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from typing import Iterable, Optional, TypedDict

import openpyxl
from charset_normalizer import from_bytes
from markdown_it import MarkdownIt

log = logging.getLogger(__name__)

# ============================================================
# Constants
# ============================================================

MAX_ROWS: int = 50_000  # XLSX/CSV row count cap (DoS guard)
MAX_SHEET_CELLS: int = 5_000_000  # XLSX sheet cell count cap
MD_SAMPLE_HEADERS: list[str] = [
    "term",
    "synonyms",
    "description",
    "example",
    "category",
]
SUPPORTED_FIELDS: set[str] = {
    "term",
    "synonyms",
    "description",
    "example",
    "category",
    "skip",
}

# 헤더명 → field 자동 매핑 키워드 (case-insensitive).
# Step 1 — STRONG: 명확한 키워드. 같은 field 에 여러 헤더 매칭 시 첫 번째만 채택.
HEADER_KEYWORDS: list[tuple[str, list[str]]] = [
    ("term", ["제목", "abbreviation", "약어", "용어명", "term"]),
    ("synonyms", ["synonym", "alias", "풀네임", "full name", "english name", "동의어"]),
    (
        "description",
        [
            "뜻",
            "의미",
            "definition",
            "description",
            "해설",
            "detail",
            "meaning",
            "설명",
        ],
    ),
    ("example", ["용도", "example", "실무", "usage", "예시"]),
    ("category", ["category", "분류", "카테고리"]),
]
# Step 2 — 헤더 "용어" / "name" 은 컨텍스트 의존:
#   - term 미할당 + 다른 헤더가 term 후보 없음 → "용어" 가 term
#   - 이미 term 있음 (예: "제목" 매칭) → "용어" 는 synonyms
_AMBIGUOUS_TERM_KEYWORDS: list[str] = ["용어", "name"]

# Category auto-classify keywords (`/tmp/convert_glossaries.py` 이식, 9 카테고리).
# 우선순위 순서. 첫 매칭이 채택된다.
CATEGORY_KEYWORDS: list[tuple[str, list[str]]] = [
    (
        "운송 서류",
        [
            "B/L",
            "Bill of Lading",
            "D/O",
            "Delivery Order",
            "Manifest",
            "Shipping Instruction",
            "Booking",
            "Waybill",
            "LOI",
            "선하증권",
            "운송장",
            "Endorsement",
            "Surrender",
            "Switch BL",
        ],
    ),
    (
        "재무/결제",
        [
            "A/R",
            "A/P",
            "L/C",
            "Letter of Credit",
            "Freight",
            "Demurrage",
            "Surcharge",
            "BAF",
            "CAF",
            "THC",
            "DCF",
            "Bank LG",
            "Bank Guarantee",
            "외상",
            "운임",
            "수수료",
            "Charge",
            "Invoice",
        ],
    ),
    (
        "컨테이너",
        [
            "Container",
            "Reefer",
            "DG",
            "Dangerous",
            "OOG",
            "FlexiBag",
            "GOH",
            "SOC",
            "BOLT SEAL",
            "Seal",
            "VGM",
            "MSDS",
            "컨테이너",
            "위험물",
            "냉동",
            "Lithium-ion",
            "Coil",
        ],
    ),
    (
        "통관/세관",
        [
            "Customs",
            "ACI",
            "CCAM",
            "SCMTR",
            "AFR",
            "MPCI",
            "Declaration",
            "통관",
            "세관",
            "신고",
            "AEO",
            "AMS",
            "Filing",
        ],
    ),
    (
        "선박/항만",
        [
            "Vessel",
            "Port",
            "Terminal",
            "Berth",
            "Voyage",
            "ETA",
            "ETD",
            "Stowage",
            "Slot",
            "Trucking",
            "선박",
            "선석",
            "부두",
            "부산",
            "인천",
            "Panamax",
            "Tanker",
        ],
    ),
    (
        "보험/책임",
        [
            "Insurance",
            "GA",
            "General Average",
            "Salvage",
            "Claim",
            "보험",
            "공동해손",
            "클레임",
            "면책",
            "책임",
        ],
    ),
    (
        "정책/규정",
        [
            "Policy",
            "Sanction",
            "IMO",
            "SOLAS",
            "MOU",
            "ISO",
            "Convention",
            "정책",
            "규정",
        ],
    ),
    (
        "회사/조직",
        [
            "Affiliated",
            "Agency",
            "Forwarder",
            "Shipper",
            "Consignee",
            "Carrier",
            "Agent",
            "관계회사",
            "포워더",
            "송하인",
            "수하인",
        ],
    ),
    (
        "화물/운영",
        [
            "Cargo",
            "Export",
            "Import",
            "T/S",
            "Transshipment",
            "Shipback",
            "COD",
            "Empty",
            "EOV",
            "수출",
            "수입",
            "환적",
        ],
    ),
]


class ParsedFile(TypedDict):
    """파싱 결과 표준 컨테이너.

    ``headers`` 는 XLSX/CSV 에서는 실제 컬럼명, MD 에서는 ``MD_SAMPLE_HEADERS`` 가 사용된다.
    """

    format: str  # "xlsx" | "csv" | "md"
    headers: list[str]
    rows: list[dict]  # 각 row 는 header → cell 값
    encoding: Optional[str]
    md_pattern: Optional[str]  # "sectioned" | "table" | "h2-dl" | None


class GlossaryParseError(ValueError):
    """파싱 단계의 사용자 가시 에러. 라우터에서 400 으로 변환."""


# ============================================================
# XLSX
# ============================================================


def parse_xlsx(file_bytes: bytes, max_rows: int = MAX_ROWS) -> ParsedFile:
    """XLSX → ParsedFile.

    - ``read_only=True, data_only=True`` 로 streaming 파싱 + formula 무시 (수식 결과만).
    - 첫 sheet 의 첫 행 = 헤더, 나머지 = data row.
    - sheet 셀 수 ``MAX_SHEET_CELLS`` 초과 시 거절.
    - row count 가 ``max_rows`` 초과 시 거절.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as e:
        raise GlossaryParseError(f"XLSX 파일을 읽을 수 없습니다: {e}") from e

    try:
        ws = wb.active
        if ws is None:
            raise GlossaryParseError("XLSX 에 시트가 없습니다.")

        # 셀 수 cap. read_only 모드에서는 ws.max_row * ws.max_column 이 추정치.
        # 일부 파일은 max_row 가 0 으로 보고됨 → iter 카운트로 검증 fallback.
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row * max_col > MAX_SHEET_CELLS:
            raise GlossaryParseError(
                f"시트가 너무 큽니다: {max_row}행 × {max_col}열 (한도 {MAX_SHEET_CELLS} 셀)."
            )

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise GlossaryParseError("XLSX 가 비어 있습니다.")

        headers = [_normalize_header(c) for c in header_row]
        # 빈 헤더 만으로는 의미 없음
        if not any(h for h in headers):
            raise GlossaryParseError("XLSX 의 첫 행이 모두 비어 있습니다 (헤더 누락).")

        rows: list[dict] = []
        for i, row in enumerate(rows_iter, start=2):
            if i - 1 > max_rows:
                raise GlossaryParseError(f"row 수가 한도({max_rows})를 초과합니다.")
            if not row or all(_is_empty(c) for c in row):
                continue
            rows.append(_row_to_dict(headers, row))

        return ParsedFile(
            format="xlsx",
            headers=headers,
            rows=rows,
            encoding=None,
            md_pattern=None,
        )
    finally:
        wb.close()


# ============================================================
# CSV
# ============================================================


_CSV_ENCODING_CHAIN: tuple[str, ...] = ("utf-8-sig", "utf-8", "cp949")


def parse_csv(file_bytes: bytes, max_rows: int = MAX_ROWS) -> ParsedFile:
    """CSV → ParsedFile.

    인코딩 fallback chain 으로 시도 후 모두 실패 시 ``charset-normalizer`` 자동 감지.
    delimiter 는 ``csv.Sniffer`` 로 감지 (실패 시 ``,`` 디폴트).
    """
    text, encoding = _decode_csv(file_bytes)
    if not text.strip():
        raise GlossaryParseError("CSV 가 비어 있습니다.")

    # delimiter 감지 — 첫 4096 chars 만 보면 충분
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:

        class _Default(csv.excel):
            delimiter = ","

        dialect = _Default()

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    try:
        header_row = next(reader)
    except StopIteration:
        raise GlossaryParseError("CSV 가 비어 있습니다.")

    headers = [_normalize_header(c) for c in header_row]
    if not any(h for h in headers):
        raise GlossaryParseError("CSV 의 첫 행이 모두 비어 있습니다 (헤더 누락).")

    rows: list[dict] = []
    for i, row in enumerate(reader, start=2):
        if i - 1 > max_rows:
            raise GlossaryParseError(f"row 수가 한도({max_rows})를 초과합니다.")
        if not row or all(_is_empty(c) for c in row):
            continue
        rows.append(_row_to_dict(headers, row))

    return ParsedFile(
        format="csv",
        headers=headers,
        rows=rows,
        encoding=encoding,
        md_pattern=None,
    )


def _decode_csv(file_bytes: bytes) -> tuple[str, str]:
    """CSV bytes → (text, encoding). 우선순위 chain → charset-normalizer fallback."""
    for enc in _CSV_ENCODING_CHAIN:
        try:
            return file_bytes.decode(enc), enc
        except UnicodeDecodeError:
            continue
    # charset-normalizer fallback
    try:
        result = from_bytes(file_bytes).best()
        if result is not None:
            return str(result), result.encoding or "auto"
    except Exception as e:
        log.warning("charset-normalizer 실패: %s", e)
    raise GlossaryParseError(
        "CSV 인코딩을 인식할 수 없습니다. UTF-8 또는 CP949 로 저장해주세요."
    )


# ============================================================
# Markdown
# ============================================================


# Sectioned 패턴 (ACME terminology.md): [N] title 용어:/뜻:/실무 예시:/해설:
_SECTIONED_ENTRY_SPLIT = re.compile(r"(?=\[\d+\]\s)")
_SECTIONED_HEADER = re.compile(r"^\[(\d+)\]\s*(.+?)\s+용어:", re.DOTALL)
_SECTIONED_LABELS = re.compile(
    r"(용어|뜻|실무 예시|해설):\s*(.*?)(?=\s+(?:용어:|뜻:|실무 예시:|해설:)|$)",
    re.DOTALL,
)
# H2/dl 패턴: ## Term ... - field: value
_H2_HEADING = re.compile(r"^##\s+(.+?)\s*$", re.MULTILINE)
_DL_FIELD = re.compile(r"^\s*-\s+([A-Za-z가-힣]+)\s*[:：]\s*(.+?)\s*$")


def parse_md(file_bytes: bytes) -> ParsedFile:
    """Markdown → ParsedFile. 감지 우선순위: table → sectioned → h2-dl."""
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            raise GlossaryParseError("Markdown 은 UTF-8 인코딩이어야 합니다.")
    text = text.strip()
    if not text:
        raise GlossaryParseError("Markdown 이 비어 있습니다.")

    # 1) Table — markdown-it 의 table token 우선
    rows = _parse_md_table(text)
    if rows is not None and rows:
        # table 은 헤더가 실제 컬럼이라 그대로 전달
        headers = list(rows[0].keys()) if rows else MD_SAMPLE_HEADERS
        return ParsedFile(
            format="md",
            headers=headers,
            rows=rows,
            encoding=None,
            md_pattern="table",
        )

    # 2) Sectioned (ACME terminology.md)
    rows = _parse_md_sectioned(text)
    if rows:
        return ParsedFile(
            format="md",
            headers=MD_SAMPLE_HEADERS,
            rows=rows,
            encoding=None,
            md_pattern="sectioned",
        )

    # 3) H2 + definition list
    rows = _parse_md_h2_dl(text)
    if rows:
        return ParsedFile(
            format="md",
            headers=MD_SAMPLE_HEADERS,
            rows=rows,
            encoding=None,
            md_pattern="h2-dl",
        )

    raise GlossaryParseError(
        "Markdown 패턴을 인식할 수 없습니다. 지원: table / sectioned ([N] 용어:/뜻:/...) / H2 + dl."
    )


def _parse_md_table(text: str) -> Optional[list[dict]]:
    """markdown-it 의 table token 으로 첫 번째 테이블 파싱."""
    md = MarkdownIt("commonmark").enable("table")
    try:
        tokens = md.parse(text)
    except Exception as e:
        log.warning("markdown-it 파싱 실패: %s", e)
        return None

    headers: list[str] = []
    rows: list[dict] = []
    in_table = False
    in_thead = False
    in_tbody = False
    current_row: list[str] = []
    in_cell = False
    cell_buf: list[str] = []

    for tok in tokens:
        if tok.type == "table_open":
            in_table = True
        elif tok.type == "table_close":
            in_table = False
        elif tok.type == "thead_open":
            in_thead = True
        elif tok.type == "thead_close":
            in_thead = False
        elif tok.type == "tbody_open":
            in_tbody = True
        elif tok.type == "tbody_close":
            in_tbody = False
        elif tok.type in ("th_open", "td_open"):
            in_cell = True
            cell_buf = []
        elif tok.type in ("th_close", "td_close"):
            cell_text = "".join(cell_buf).strip()
            if in_thead:
                headers.append(cell_text)
            else:
                current_row.append(cell_text)
            in_cell = False
        elif tok.type == "tr_open":
            current_row = []
        elif tok.type == "tr_close":
            if in_tbody and current_row:
                row = _row_to_dict([_normalize_header(h) for h in headers], current_row)
                if any(v for v in row.values()):
                    rows.append(row)
        elif in_cell and tok.type == "inline":
            # inline 자식에서 텍스트 모음
            if tok.children:
                for child in tok.children:
                    if child.type == "text":
                        cell_buf.append(child.content)
                    elif child.type == "softbreak" or child.type == "hardbreak":
                        cell_buf.append("\n")
                    elif child.type == "code_inline":
                        cell_buf.append(child.content)

        if not in_table and rows:
            # 첫 테이블 끝나면 stop
            break

    if not headers:
        return None
    return rows


_SECTIONED_PIECE_MAX = (
    16_384  # piece 단위 길이 cap — ReDoS 방어 (DOTALL + .+? 백트래킹)
)


def _parse_md_sectioned(text: str) -> list[dict]:
    """ACME terminology.md 형식 파싱."""
    # 분리자 제거
    body = re.sub(r"-{10,}", " ", text)
    body = re.sub(r"={10,}", " ", body)
    pieces = [p.strip() for p in _SECTIONED_ENTRY_SPLIT.split(body) if p.strip()]
    rows: list[dict] = []
    for piece in pieces:
        # ReDoS 방어: 비정상적으로 큰 piece (단일 [N] 뒤 거대한 본문) 는 prefix 만 검사.
        # 한 entry 가 16KB 넘는 정상 글로서리는 없음.
        if len(piece) > _SECTIONED_PIECE_MAX:
            piece = piece[:_SECTIONED_PIECE_MAX]
        m = _SECTIONED_HEADER.match(piece)
        if not m:
            continue
        main_title = m.group(2).strip()
        sections: dict[str, str] = {}
        for sm in _SECTIONED_LABELS.finditer(piece):
            sections[sm.group(1)] = sm.group(2).strip()

        description = sections.get("뜻", "").strip()
        detail = sections.get("해설", "").strip()
        if detail:
            detail_fmt = re.sub(r"\s*□\s*", "\n• ", detail).strip()
            description = f"{description}\n\n[해설]\n{detail_fmt}".strip()

        rows.append(
            {
                "term": main_title,
                "synonyms": sections.get("용어", main_title),
                "description": description,
                "example": sections.get("실무 예시", ""),
                "category": "",
            }
        )
    return rows


def _parse_md_h2_dl(text: str) -> list[dict]:
    """## Term ... - field: value 패턴."""
    rows: list[dict] = []
    # H2 위치로 분할
    blocks = re.split(r"(?=^##\s+)", text, flags=re.MULTILINE)
    for block in blocks:
        block = block.strip()
        if not block.startswith("##"):
            continue
        # 첫 줄 = ## Term
        lines = block.splitlines()
        head_match = _H2_HEADING.match(lines[0])
        if not head_match:
            continue
        term = head_match.group(1).strip()
        if not term:
            continue
        fields: dict[str, str] = {"term": term}
        for line in lines[1:]:
            fm = _DL_FIELD.match(line)
            if not fm:
                continue
            key = _field_alias(fm.group(1).strip())
            if key:
                fields.setdefault(key, fm.group(2).strip())
        rows.append(fields)
    return rows


_H2_FIELD_ALIASES: dict[str, str] = {
    "synonyms": "synonyms",
    "synonym": "synonyms",
    "alias": "synonyms",
    "동의어": "synonyms",
    "description": "description",
    "definition": "description",
    "뜻": "description",
    "의미": "description",
    "설명": "description",
    "example": "example",
    "예시": "example",
    "용도": "example",
    "category": "category",
    "분류": "category",
    "카테고리": "category",
}


def _field_alias(raw: str) -> Optional[str]:
    return _H2_FIELD_ALIASES.get(raw.lower()) or _H2_FIELD_ALIASES.get(raw)


# ============================================================
# Header inference
# ============================================================


def infer_header_mapping(headers: Iterable[str]) -> dict[str, str]:
    """headers → {header_name: field_name | "skip"}.

    field_name ∈ {term, synonyms, description, example, category}.

    Step 1: STRONG 키워드 매칭. 동일 field 매칭 헤더가 여러 개면 첫 번째만 채택.
    Step 2: "용어"/"name" 헤더는 컨텍스트 의존 — term 미할당이면 term, 아니면 synonyms.
    Step 3: 미매칭 헤더는 skip.
    Step 4: term 이 끝까지 없으면 첫 비어있지 않은 헤더를 term 으로 force-fallback.
    """
    mapping: dict[str, str] = {}
    used_fields: set[str] = set()
    header_list = list(headers)

    # Step 1: STRONG 매칭
    for field, keywords in HEADER_KEYWORDS:
        if field in used_fields:
            continue
        for h in header_list:
            if h in mapping:
                continue
            if _header_matches(h, keywords):
                mapping[h] = field
                used_fields.add(field)
                break

    # Step 2: "용어"/"name" 컨텍스트 매핑
    for h in header_list:
        if h in mapping or not h:
            continue
        if _header_matches(h, _AMBIGUOUS_TERM_KEYWORDS):
            if "term" not in used_fields:
                mapping[h] = "term"
                used_fields.add("term")
            elif "synonyms" not in used_fields:
                mapping[h] = "synonyms"
                used_fields.add("synonyms")

    # Step 3: 미매칭 헤더 skip
    for h in header_list:
        if h and h not in mapping:
            mapping[h] = "skip"

    # Step 4: term force-fallback. mapping 이 skip 이거나 미할당 인 헤더만 대상 —
    # 다른 field 에 이미 매핑된 헤더를 덮어쓰지 않는다.
    if "term" not in used_fields:
        for h in header_list:
            if h and mapping.get(h, "skip") == "skip":
                mapping[h] = "term"
                break

    return mapping


def _header_matches(header: str, keywords: list[str]) -> bool:
    """헤더명 ↔ 키워드 case-insensitive contains 매칭."""
    if not header:
        return False
    h_lower = header.lower()
    for kw in keywords:
        if kw.lower() in h_lower:
            return True
    return False


# ============================================================
# Normalization
# ============================================================


_SYNONYM_SPLIT = re.compile(r"[,、]")


def normalize_entries(
    parsed: ParsedFile,
    mapping: dict[str, str],
    auto_classify: bool = False,
) -> tuple[list[dict], dict[str, int]]:
    """ParsedFile → glossary entry dict list + 통계.

    - ``mapping`` 의 값이 'skip' 인 헤더는 무시.
    - 파일내 중복 term 은 first wins. 이후 중복 row 는 skip 카운트.
    - 빈 term row 는 skip.
    - synonyms 는 ``,`` ``、`` split. ``/`` 보존. primary term 제거.
    - description 의 ``□`` → ``\\n• ``.
    - ``auto_classify=True`` 이면 category 가 비어있는 entry 에 keyword 분류 적용.

    Returns: (entries, stats {"total", "parsed", "skipped"})
    """
    entries: list[dict] = []
    seen_terms: dict[str, int] = {}  # lowercase term → index in entries
    skipped = 0
    rows = parsed["rows"]

    for row in rows:
        fields: dict[str, str] = {
            "term": "",
            "synonyms": "",
            "description": "",
            "example": "",
            "category": "",
        }
        for header, field in mapping.items():
            if field == "skip" or field not in fields:
                continue
            value = row.get(header, "")
            if value is None:
                continue
            value = str(value).strip()
            if not value:
                continue
            # 같은 field 에 이미 값이 있으면 (header 가 여러 개 같은 field 로 매핑된 케이스 — 통상 발생 안 함)
            # 첫 값 유지
            if not fields[field]:
                fields[field] = value

        term = fields["term"].strip()
        if not term:
            skipped += 1
            continue

        # 파일내 중복 term — first wins
        key = term.lower()
        if key in seen_terms:
            skipped += 1
            continue

        # synonyms split (콤마/한국어 콤마. 슬래시 보존.)
        synonyms = [
            s.strip() for s in _SYNONYM_SPLIT.split(fields["synonyms"]) if s.strip()
        ]
        # primary term dedupe
        synonyms = [s for s in synonyms if s != term]

        # bullet 정규화 (ACME legacy: □ → \n• )
        description = fields["description"]
        if "□" in description:
            description = re.sub(r"\s*□\s*", "\n• ", description).strip()

        category = fields["category"].strip()
        if not category and auto_classify:
            category = auto_classify_category(term, synonyms, description) or ""

        entry: dict = {
            "term": term,
            "synonyms": synonyms,
            "description": description,
            "example": fields["example"],
        }
        if category:
            entry["category"] = category

        seen_terms[key] = len(entries)
        entries.append(entry)

    stats = {
        "total": len(rows),
        "parsed": len(entries),
        "skipped": skipped,
    }
    return entries, stats


# ============================================================
# Auto-classify
# ============================================================


def auto_classify_category(
    term: str,
    synonyms: list[str],
    description: str = "",
) -> Optional[str]:
    """Keyword 기반 카테고리 분류.

    1) term + synonyms (정확 매칭 우선)
    2) description 첫 500 chars (fallback)
    매칭 없으면 None.

    word boundary regex 로 false positive 방지 — 예: ``BL`` 이 ``receivABLe`` 매칭 X.
    """
    primary = f" {term} " + " ".join(f" {s} " for s in synonyms)
    secondary = (description or "")[:500]

    for category, keywords in CATEGORY_KEYWORDS:
        for kw in keywords:
            if _keyword_match(kw, primary):
                return category

    for category, keywords in CATEGORY_KEYWORDS:
        for kw in keywords:
            if _keyword_match(kw, secondary):
                return category

    return None


def _keyword_match(kw: str, text: str) -> bool:
    """Word boundary 매칭.

    - 슬래시/점/짧은 대문자 약어 (≤4글자): case-sensitive + 영문자 인접 X.
    - 그 외: case-insensitive ``\\b...\\b``.
    """
    if "/" in kw or "." in kw or (any(c.isupper() for c in kw) and len(kw) <= 4):
        pattern = r"(?<![A-Za-z])" + re.escape(kw) + r"(?![A-Za-z])"
        return re.search(pattern, text) is not None
    pattern = r"\b" + re.escape(kw) + r"\b"
    return re.search(pattern, text, re.IGNORECASE) is not None


# ============================================================
# Helpers
# ============================================================


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _row_to_dict(headers: list[str], row: Iterable[object]) -> dict:
    out: dict[str, str] = {}
    for header, cell in zip(headers, row):
        if not header:
            continue
        if cell is None:
            out[header] = ""
        else:
            out[header] = str(cell).strip() if isinstance(cell, str) else str(cell)
    return out


# ============================================================
# Public entry — file extension dispatcher
# ============================================================


def parse_file(
    filename: str,
    file_bytes: bytes,
    max_rows: int = MAX_ROWS,
) -> ParsedFile:
    """확장자 기반 dispatcher. router 가 단일 진입점으로 호출."""
    if not filename:
        raise GlossaryParseError("파일명이 없습니다.")
    name = filename.lower()
    if name.endswith(".xlsx"):
        return parse_xlsx(file_bytes, max_rows=max_rows)
    if name.endswith(".csv"):
        return parse_csv(file_bytes, max_rows=max_rows)
    if name.endswith(".md") or name.endswith(".markdown"):
        return parse_md(file_bytes)
    raise GlossaryParseError(
        f"지원하지 않는 파일 확장자: {filename} (지원: .xlsx, .csv, .md)"
    )
