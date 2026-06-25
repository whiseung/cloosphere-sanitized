import logging
import os
import sys
import tempfile
from typing import Optional

import ftfy
import requests
from langchain_community.document_loaders import (
    AzureAIDocumentIntelligenceLoader,
    BSHTMLLoader,
    CSVLoader,
    Docx2txtLoader,
    PyPDFLoader,
    TextLoader,
    UnstructuredEmailLoader,
    UnstructuredEPubLoader,
    UnstructuredExcelLoader,
    UnstructuredPowerPointLoader,
    UnstructuredRSTLoader,
    UnstructuredXMLLoader,
)
from langchain_core.documents import Document
from open_webui.env import GLOBAL_LOG_LEVEL, SRC_LOG_LEVELS
from open_webui.retrieval.loaders.mistral import MistralLoader

logging.basicConfig(stream=sys.stdout, level=GLOBAL_LOG_LEVEL)
log = logging.getLogger(__name__)
log.setLevel(SRC_LOG_LEVELS["RAG"])

known_source_ext = [
    "go",
    "py",
    "java",
    "sh",
    "bat",
    "ps1",
    "cmd",
    "js",
    "ts",
    "css",
    "cpp",
    "hpp",
    "h",
    "c",
    "cs",
    "sql",
    "log",
    "ini",
    "pl",
    "pm",
    "r",
    "dart",
    "dockerfile",
    "env",
    "php",
    "hs",
    "hsc",
    "lua",
    "nginxconf",
    "conf",
    "m",
    "mm",
    "plsql",
    "perl",
    "rb",
    "rs",
    "db2",
    "scala",
    "bash",
    "swift",
    "vue",
    "svelte",
    "msg",
    "ex",
    "exs",
    "erl",
    "tsx",
    "jsx",
    "hs",
    "lhs",
    "json",
]


def _flatten_xlsx_merges(src_path: str) -> Optional[str]:
    """openpyxl 로 xlsx 의 모든 병합 셀을 unmerge 하고 좌상단 값을 해당 영역
    전체에 복제한 뒤 임시 파일로 저장. 그 임시 파일을 unstructured 에 넘기면
    pandas 가 평탄화해도 셀 값이 모든 위치에 살아 있어 HTML <td> 가 빈 칸으로
    떨어지지 않는다.

    .xls (legacy binary) 는 openpyxl 이 못 읽으므로 None 반환 — 호출부가
    원본 파일을 그대로 사용한다.
    """
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(src_path, data_only=True)
        for ws in wb.worksheets:
            for merged in list(ws.merged_cells.ranges):
                top_left = ws.cell(merged.min_row, merged.min_col).value
                ws.unmerge_cells(str(merged))
                for r in range(merged.min_row, merged.max_row + 1):
                    for c in range(merged.min_col, merged.max_col + 1):
                        ws.cell(r, c).value = top_left
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        return tmp.name
    except Exception as e:
        log.warning("Failed to flatten xlsx merges: %s", e)
        return None


class TikaLoader:
    def __init__(self, url, file_path, mime_type=None):
        self.url = url
        self.file_path = file_path
        self.mime_type = mime_type

    def load(self) -> list[Document]:
        with open(self.file_path, "rb") as f:
            data = f.read()

        if self.mime_type is not None:
            headers = {"Content-Type": self.mime_type}
        else:
            headers = {}

        endpoint = self.url
        if not endpoint.endswith("/"):
            endpoint += "/"
        endpoint += "tika/text"

        r = requests.put(endpoint, data=data, headers=headers)

        if r.ok:
            raw_metadata = r.json()
            text = raw_metadata.get("X-TIKA:content", "<No text content found>").strip()

            if "Content-Type" in raw_metadata:
                headers["Content-Type"] = raw_metadata["Content-Type"]

            log.debug("Tika extracted text: %s", text)

            return [Document(page_content=text, metadata=headers)]
        else:
            raise Exception(f"Error calling Tika: {r.reason}")


class DoclingLoader:
    def __init__(self, url, file_path=None, mime_type=None):
        self.url = url.rstrip("/")
        self.file_path = file_path
        self.mime_type = mime_type

    def load(self) -> list[Document]:
        with open(self.file_path, "rb") as f:
            files = {
                "files": (
                    self.file_path,
                    f,
                    self.mime_type or "application/octet-stream",
                )
            }

            params = {
                "image_export_mode": "placeholder",
                "table_mode": "accurate",
                "include_images": "false",
                "do_picture_description": "false",
                "do_picture_classification": "false",
                "force_ocr": "false",
            }

            # docling-serve >= 0.5.0 uses /v1, older versions used /v1alpha
            endpoint = f"{self.url}/v1/convert/file"
            r = requests.post(endpoint, files=files, data=params)
            # Fallback for older docling-serve versions
            if r.status_code == 404:
                endpoint = f"{self.url}/v1alpha/convert/file"
                with open(self.file_path, "rb") as f2:
                    files2 = {
                        "files": (
                            self.file_path,
                            f2,
                            self.mime_type or "application/octet-stream",
                        )
                    }
                    r = requests.post(endpoint, files=files2, data=params)

        if r.ok:
            result = r.json()
            document_data = result.get("document", {})
            text = document_data.get("md_content", "<No text content found>")

            metadata = {"Content-Type": self.mime_type} if self.mime_type else {}

            log.debug("Docling extracted text: %s", text)

            return [Document(page_content=text, metadata=metadata)]
        else:
            error_msg = f"Error calling Docling API: {r.reason}"
            if r.text:
                try:
                    error_data = r.json()
                    if "detail" in error_data:
                        error_msg += f" - {error_data['detail']}"
                except Exception:
                    error_msg += f" - {r.text}"
            raise Exception(f"Error calling Docling: {error_msg}")


class Loader:
    def __init__(self, engine: str = "", **kwargs):
        self.engine = engine
        self.kwargs = kwargs
        self._temp_files: list[str] = []

    def load(
        self, filename: str, file_content_type: str, file_path: str
    ) -> list[Document]:
        try:
            loader = self._get_loader(filename, file_content_type, file_path)
            docs = loader.load()

            # 사전 처리(xlsx 병합 평탄화)로 임시 파일을 unstructured 에 넘긴 경우,
            # element.metadata.{source,filename,file_path} 가 임시 경로/이름으로
            # 박혀 있을 수 있어 원본 file_path 로 덮어쓴다 — 다운스트림 표시/소스
            # 추적용. 전체 경로와 basename 둘 다 매칭.
            tmp_full = set(self._temp_files)
            tmp_base = {os.path.basename(p) for p in self._temp_files}

            out: list[Document] = []
            for doc in docs:
                text = doc.page_content
                # Table element 처럼 구조가 있는 청크는 unstructured 가
                # metadata.text_as_html 에 HTML <table> 을 채워둔다. RAG 청크에서
                # 표 구조 보존을 위해 평문 대신 HTML 을 page_content 로 promote.
                html = (doc.metadata or {}).get("text_as_html")
                if html:
                    text = html
                metadata = dict(doc.metadata or {})
                for key in ("source", "file_path"):
                    if metadata.get(key) in tmp_full:
                        metadata[key] = file_path
                if metadata.get("filename") in tmp_base:
                    metadata["filename"] = filename
                out.append(
                    Document(page_content=ftfy.fix_text(text), metadata=metadata)
                )
            return out
        finally:
            # 사전 처리에서 생성한 임시 파일 정리 (xlsx 병합 평탄화 등).
            for tmp in self._temp_files:
                try:
                    os.unlink(tmp)
                except OSError:
                    pass
            self._temp_files = []

    def _is_text_file(self, file_ext: str, file_content_type: str) -> bool:
        return file_ext in known_source_ext or (
            file_content_type and file_content_type.find("text/") >= 0
        )

    def _get_loader(self, filename: str, file_content_type: str, file_path: str):
        file_ext = filename.split(".")[-1].lower()

        # 엔진이 지정됐는데 필수 자격증명이 비어 있으면 조용히 fallback 하지 말고 명시적 실패.
        # (fallback 시 PyPDFLoader 같은 내장 로더가 동작해 사용자가 원했던 외부 엔진을 우회한 줄 모름)
        if self.engine == "tika" and not self.kwargs.get("TIKA_SERVER_URL"):
            raise ValueError(
                "Content extraction engine is 'tika' but TIKA_SERVER_URL is not configured."
            )
        if self.engine == "docling" and not self.kwargs.get("DOCLING_SERVER_URL"):
            raise ValueError(
                "Content extraction engine is 'docling' but DOCLING_SERVER_URL is not configured."
            )
        if self.engine == "document_intelligence" and (
            not self.kwargs.get("DOCUMENT_INTELLIGENCE_ENDPOINT")
            or not self.kwargs.get("DOCUMENT_INTELLIGENCE_KEY")
        ):
            raise ValueError(
                "Content extraction engine is 'document_intelligence' but endpoint/key is not configured."
            )
        if self.engine == "mistral_ocr" and not self.kwargs.get("MISTRAL_OCR_API_KEY"):
            raise ValueError(
                "Content extraction engine is 'mistral_ocr' but MISTRAL_OCR_API_KEY is not configured."
            )
        if self.engine == "document_ai" and not self.kwargs.get(
            "DOCUMENT_AI_PROCESSOR_ID"
        ):
            raise ValueError(
                "Content extraction engine is 'document_ai' but processor ID is not configured."
            )

        if self.engine == "tika" and self.kwargs.get("TIKA_SERVER_URL"):
            if self._is_text_file(file_ext, file_content_type):
                loader = TextLoader(file_path, autodetect_encoding=True)
            else:
                loader = TikaLoader(
                    url=self.kwargs.get("TIKA_SERVER_URL"),
                    file_path=file_path,
                    mime_type=file_content_type,
                )
        elif self.engine == "docling" and self.kwargs.get("DOCLING_SERVER_URL"):
            if self._is_text_file(file_ext, file_content_type):
                loader = TextLoader(file_path, autodetect_encoding=True)
            else:
                loader = DoclingLoader(
                    url=self.kwargs.get("DOCLING_SERVER_URL"),
                    file_path=file_path,
                    mime_type=file_content_type,
                )
        elif (
            self.engine == "document_intelligence"
            and self.kwargs.get("DOCUMENT_INTELLIGENCE_ENDPOINT") != ""
            and self.kwargs.get("DOCUMENT_INTELLIGENCE_KEY") != ""
            and (
                file_ext in ["pdf", "xls", "xlsx", "docx", "ppt", "pptx"]
                or file_content_type
                in [
                    "application/vnd.ms-excel",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    "application/vnd.ms-powerpoint",
                    "application/vnd.openxmlformats-officedocument.presentationml.presentation",
                ]
            )
        ):
            loader = AzureAIDocumentIntelligenceLoader(
                file_path=file_path,
                api_endpoint=self.kwargs.get("DOCUMENT_INTELLIGENCE_ENDPOINT"),
                api_key=self.kwargs.get("DOCUMENT_INTELLIGENCE_KEY"),
                # 고해상도 OCR — 조밀한 표/소형 폰트(데이터시트·도면 등) 인식률 향상.
                # 문서처리 프로파일의 추출엔진 config 에서 켜면 전달된다(기본 off).
                **(
                    {"analysis_features": ["ocrHighResolution"]}
                    if self.kwargs.get("DOCUMENT_INTELLIGENCE_HIGH_RESOLUTION")
                    else {}
                ),
            )
        elif (
            self.engine == "mistral_ocr"
            and self.kwargs.get("MISTRAL_OCR_API_KEY") != ""
            and file_ext
            in ["pdf"]  # Mistral OCR currently only supports PDF and images
        ):
            loader = MistralLoader(
                api_key=self.kwargs.get("MISTRAL_OCR_API_KEY"), file_path=file_path
            )
        elif (
            self.engine == "document_ai"
            and self.kwargs.get("DOCUMENT_AI_PROCESSOR_ID") != ""
            and file_ext
            in [
                "pdf",
                "tiff",
                "gif",
                "jpg",
                "jpeg",
                "png",
                "bmp",
                "webp",
                "docx",
                "xlsx",
                "pptx",
            ]
        ):
            from open_webui.retrieval.loaders.google_document_ai import (
                GoogleDocumentAILoader,
            )

            loader = GoogleDocumentAILoader(
                project_id=self.kwargs.get("DOCUMENT_AI_PROJECT_ID"),
                location=self.kwargs.get("DOCUMENT_AI_LOCATION", "us"),
                processor_id=self.kwargs.get("DOCUMENT_AI_PROCESSOR_ID"),
                processor_version=self.kwargs.get("DOCUMENT_AI_PROCESSOR_VERSION")
                or "",
                service_account_key=self.kwargs.get("DOCUMENT_AI_SERVICE_ACCOUNT_KEY")
                or self.kwargs.get("GOOGLE_CLOUD_SERVICE_ACCOUNT_KEY")
                or "",
                file_path=file_path,
                filename=filename,  # 원본 파일명 전달 (MIME 타입 감지용)
            )
        else:
            if file_ext == "pdf":
                loader = PyPDFLoader(
                    file_path, extract_images=self.kwargs.get("PDF_EXTRACT_IMAGES")
                )
            elif file_ext == "csv":
                loader = CSVLoader(file_path, autodetect_encoding=True)
            elif file_ext == "rst":
                loader = UnstructuredRSTLoader(file_path, mode="elements")
            elif file_ext == "xml":
                loader = UnstructuredXMLLoader(file_path)
            elif file_ext in ["htm", "html"]:
                loader = BSHTMLLoader(file_path, open_encoding="unicode_escape")
            elif file_ext == "md":
                loader = TextLoader(file_path, autodetect_encoding=True)
            elif file_content_type == "application/epub+zip":
                loader = UnstructuredEPubLoader(file_path)
            elif (
                file_content_type
                == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                or file_ext == "docx"
            ):
                loader = Docx2txtLoader(file_path)
            elif file_content_type in [
                "application/vnd.ms-excel",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ] or file_ext in ["xls", "xlsx"]:
                # 1) openpyxl 로 병합 셀을 unmerge + 좌상단 값 복제 (pandas 가
                #    이후 평탄화해도 모든 셀에 값이 살아남도록).
                # 2) mode="elements": 시트별 표를 별도 Document 로 분리하고
                #    text_as_html 메타 유지 → Loader.load() promote 로직이 HTML
                #    표를 청크로 들여보냄.
                # 3) include_header=True: 컬럼 헤더 행을 포함.
                target_path = file_path
                if file_ext == "xlsx":
                    flat = _flatten_xlsx_merges(file_path)
                    if flat:
                        target_path = flat
                        self._temp_files.append(flat)
                loader = UnstructuredExcelLoader(
                    target_path, mode="elements", include_header=True
                )
            elif file_content_type in [
                "application/vnd.ms-powerpoint",
                "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            ] or file_ext in ["ppt", "pptx"]:
                loader = UnstructuredPowerPointLoader(file_path)
            elif file_ext in ["eml", "msg"]:
                # 이메일(.eml MIME / .msg Outlook) 본문만 추출. 첨부 파싱 금지
                # (process_attachments=False) — 첨부 base64/파일명이 본문에 섞여
                # KB 에 색인되는 유출 방지. 이전 기본 경로(TextLoader raw MIME)는
                # 첨부를 그대로 노출했으므로 이 분기가 보안상 개선이다.
                # .msg 는 unstructured 가 oxmsg(python-oxmsg) 로 파싱 → extract_msg
                # 불필요. 파싱 실패 시 raw 폴백을 두지 않는다(폴백이 첨부 유출·OLE
                # mojibake 를 재유발) → 파일별 에러로 깨끗이 실패.
                #
                # 순서 불변식: .msg 는 known_source_ext(L74)에도 있으므로 이 분기는
                # 아래 _is_text_file 분기보다 반드시 먼저 와야 한다(아니면 .msg 가
                # TextLoader 로 샘). 기본 추출 엔진 한정 — tika/docling 선택 시엔
                # .eml 만 해당 엔진이 처리하고, .msg 는 _is_text_file 에서 TextLoader
                # 로 빠진다(기존 동작, 본 변경 범위 밖).
                loader = UnstructuredEmailLoader(file_path, process_attachments=False)
            elif self._is_text_file(file_ext, file_content_type):
                loader = TextLoader(file_path, autodetect_encoding=True)
            else:
                loader = TextLoader(file_path, autodetect_encoding=True)

        return loader
