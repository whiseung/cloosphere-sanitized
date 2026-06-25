"""create_xlsx — Excel 스프레드시트 생성 툴."""

import logging
import re
from io import BytesIO
from typing import Optional, Union

from extension_modules.tools.document._common import (
    SYSTEM_TOKEN_KEYS,
    TOKEN_PATTERN,
    format_tool_response,
    load_xlsx_template,
    save_to_files,
    system_token_values,
)
from langchain_core.tools import StructuredTool
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, model_validator

log = logging.getLogger(__name__)

CellValue = Union[str, int, float, None]

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ── Styling 상수 ──────────────────────────────────────────────────────────────
_THIN = Side(border_style="thin", color="000000")
_MEDIUM = Side(border_style="medium", color="000000")
_DATA_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
# 헤더 행: 본문과 구분을 위해 아래쪽만 medium
_HEADER_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_MEDIUM)
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")  # 옅은 회색

_TITLE_FONT = Font(bold=True, size=14)
_TITLE_ALIGN = Alignment(horizontal="center", vertical="center")
_TITLE_ROW_HEIGHT = 24

# 레이아웃: row 1 = 제목, row 2 = 공백, row 3 = 헤더, row 4~ = 데이터
_HEADER_ROW = 3
_DATA_START_ROW = 4

_COL_WIDTH_MIN = 10
_COL_WIDTH_MAX = 50
_COL_WIDTH_PADDING = 2  # 셀 내용 끝의 여유 공간

# CSV/Formula Injection 차단 — Excel 이 셀 값을 수식으로 평가하는 prefix.
# LLM 이 생성한 untrusted 문자열을 그대로 셀에 쓰면 `=WEBSERVICE(...)` 같은 외부
# 호출 / 자격증명 송신이 가능. `'` (apostrophe) prefix 가 Excel 표준 이스케이프.
# Tab/CR 은 Excel 이 명시 수식 trigger 는 아니지만 일부 환경에서 셀 구분 깨짐 보고.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell_value(value):
    """Excel 셀에 안전하게 쓸 값으로 변환 — formula injection prefix 이스케이프.

    문자열만 검사 (숫자/None 은 그대로). prefix 발견 시 `'` 추가 — Excel 이 '를
    표시 안 하고 텍스트로만 취급. 정상 데이터에 prefix 가 우연히 들어간 케이스
    (음수 `-100`, 회계 `+5.2%`) 도 함께 escape — 보안 우선, false-positive 의
    cosmetic 비용 < 잠재 공격 비용.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _visual_width(value: CellValue) -> int:
    """문자열의 시각적 너비 — 한글/한자/일본어는 2, ASCII 는 1.

    Excel 의 column width 는 표준 폰트 너비 단위라 동아시아 문자가 ASCII 의
    약 2배 너비로 렌더링됨. autosize 시 한국어 비중을 고려해야 깨지지 않음.
    """
    s = "" if value is None else str(value)
    return sum(2 if ord(c) > 0x2E80 else 1 for c in s)


class XlsxSheet(BaseModel):
    name: str = Field(..., max_length=31, description="시트명 (Excel 제한 31자)")
    title: Optional[str] = Field(
        None,
        max_length=200,
        description="시트 1행 A1에 표시되는 제목. 미지정 시 시트명을 사용",
    )
    columns: list[str] = Field(..., min_length=1, max_length=100, description="헤더 행")
    rows: list[list[CellValue]] = Field(
        default_factory=list,
        max_length=5000,
        description="데이터 행 (각 행은 columns 길이와 일치)",
    )

    @model_validator(mode="after")
    def _validate_row_width(self):
        for i, row in enumerate(self.rows):
            if len(row) != len(self.columns):
                raise ValueError(
                    f"row {i}: 컬럼 수 불일치 (expected {len(self.columns)}, got {len(row)})"
                )
        return self


class XlsxContent(BaseModel):
    filename: str = Field(..., description="확장자 제외, 한글/영문 가능")
    template_tokens: Optional[dict[str, str]] = Field(
        None,
        description=(
            "관리자가 업로드한 템플릿의 표지/요약 시트에 `{{key}}` 형태로 적어둔 "
            "자리표시자를 런타임에 치환할 값. 예: "
            "`{'title': '2026 Q1 매출 보고서', 'date': '2026-05-20', 'author': '홍길동'}`. "
            "템플릿 미사용 시 무시되며, 키가 누락되면 `{{key}}` 가 그대로 남는다. "
            "치환은 템플릿 시트에만 적용되고, 아래 sheets 의 데이터는 영향받지 않는다."
        ),
    )
    sheets: list[XlsxSheet] = Field(..., min_length=1, max_length=20)


# ── Builder ──────────────────────────────────────────────────────────────────


def _collect_template_tokens(wb) -> set[str]:
    """워크북의 모든 문자열 셀에서 발견된 `{{key}}` 키들의 set 반환.

    빌드 시 호출해 admin 이 어떤 토큰을 박았는지 진단 로깅, tool description 동적
    주입 등에 사용. 시트 순회 비용은 표지 1장 + 데이터 시트 몇 장 수준이라 OK.
    """
    keys: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "{{" in v:
                    keys.update(TOKEN_PATTERN.findall(v))
    return keys


def _discover_template_token_keys() -> set[str]:
    """현재 활성 템플릿을 로드해 발견된 모든 토큰 키 set 반환.

    템플릿 미설정 / 파싱 실패 / IO 에러는 모두 empty set 으로 swallow — tool
    초기화 경로에서 어떤 실패도 LLM 호출 자체를 막으면 안 됨.
    """
    try:
        wb = load_xlsx_template()
        if wb is None:
            return set()
        return _collect_template_tokens(wb)
    except Exception as e:  # noqa: BLE001 — 진단 보조 함수, 절대 throw 금지
        log.warning("Failed to discover template tokens: %s", e)
        return set()


def _substitute_template_tokens(wb, tokens: dict[str, str]) -> None:
    """템플릿 워크북의 모든 문자열 셀에서 `{{key}}` 를 ``tokens[key]`` 로 치환.

    LLM 데이터 시트가 ``wb`` 에 추가되기 **전에** 호출돼야 한다 — 본문 셀의
    합법적인 ``{{...}}`` 문자열이 의도치 않게 변형되는 사고 방지.

    누락된 키는 원본 ``{{key}}`` 를 유지 (silent empty 방지 — admin 이 템플릿
    오타를 빠르게 발견하도록).
    """
    if not tokens:
        return

    def _replace(match: re.Match) -> str:
        return tokens.get(match.group(1), match.group(0))

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "{{" in v:
                    new_v = TOKEN_PATTERN.sub(_replace, v)
                    if new_v != v:
                        # 치환 결과가 LLM 출처라 formula injection 대상
                        cell.value = _safe_cell_value(new_v)


def _build_xlsx(
    content: XlsxContent,
    *,
    system_tokens: Optional[dict[str, str]] = None,
) -> BytesIO:
    """XlsxContent → in-memory .xlsx BytesIO.

    레이아웃:
    - row 1: 제목 (A1, 14pt bold, 가로 병합 + 가운데 정렬)
    - row 2: 공백 (시각적 여백)
    - row 3: 헤더
    - row 4~: 데이터

    스타일링:
    - 헤더 행: bold + 옅은 회색 배경 + 아래쪽 medium border
    - 데이터 셀: thin border 4면
    - 컬럼 너비: 헤더 + 모든 데이터 셀의 시각적 너비 max 기준 auto-fit
      (한글 2x weight, min 10, max 50). 제목은 너비 계산에서 제외 — 병합 셀이라
      개별 컬럼 폭에 영향 주면 안 됨.

    템플릿 모드:
    - 마스터 .xlsx 의 기존 시트는 보존하고 우리 시트는 뒤에 append
    - 헤더 fill (옅은 회색 hard-coded) 은 skip 해 마스터 cell style 에 맡김
    - openpyxl 이 시트명 충돌 시 자동으로 _1 등 suffix 처리
    """
    template = load_xlsx_template()
    template_used = template is not None
    if template_used:
        wb = template
        # 시스템 자동 토큰 (date/author/...) + LLM 전달 토큰 머지 — LLM 이 우선.
        # LLM 시트 append 전에 치환 — 새 데이터 시트의 합법적인 {{...}} 문자열까지
        # 건드리는 것을 차단.
        merged_tokens: dict[str, str] = {
            **(system_tokens or {}),
            **(content.template_tokens or {}),
        }
        if merged_tokens:
            discovered = _collect_template_tokens(wb)
            unfilled = discovered - merged_tokens.keys()
            if unfilled:
                log.info(
                    "create_xlsx template has unfilled tokens %s — left as-is",
                    sorted(unfilled),
                )
            _substitute_template_tokens(wb, merged_tokens)
    else:
        wb = Workbook()
        # 기본 시트 제거 — 새로 만든다
        wb.remove(wb.active)

    for sheet in content.sheets:
        ws = wb.create_sheet(title=sheet.name)
        n_cols = len(sheet.columns)

        # row 1: 제목 (병합)
        title_text = sheet.title or sheet.name
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _TITLE_ALIGN
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.row_dimensions[1].height = _TITLE_ROW_HEIGHT

        # row 3: 헤더 — fill 은 템플릿 미사용 시만 (마스터 cell style 보존)
        for col_idx, col_name in enumerate(sheet.columns, start=1):
            cell = ws.cell(row=_HEADER_ROW, column=col_idx, value=col_name)
            cell.font = _HEADER_FONT
            if not template_used:
                cell.fill = _HEADER_FILL
            cell.border = _HEADER_BORDER

        # row 4~: 데이터 (LLM 출처라 formula injection escape)
        for r_offset, row in enumerate(sheet.rows):
            row_idx = _DATA_START_ROW + r_offset
            for c_offset, value in enumerate(row):
                cell = ws.cell(
                    row=row_idx, column=c_offset + 1, value=_safe_cell_value(value)
                )
                cell.border = _DATA_BORDER

        # 컬럼 너비 auto-fit — 헤더 + 데이터에서 시각적 max 너비 (제목 제외)
        for col_idx, col_name in enumerate(sheet.columns, start=1):
            max_w = _visual_width(col_name)
            for row in sheet.rows:
                max_w = max(max_w, _visual_width(row[col_idx - 1]))
            width = max(_COL_WIDTH_MIN, min(_COL_WIDTH_MAX, max_w + _COL_WIDTH_PADDING))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


_TOOL_DESCRIPTION = (
    "Generate an Excel spreadsheet (.xlsx) file. "
    "표 데이터/스프레드시트를 .xlsx 파일로 생성한다. "
    "Use when the user asks for: 'Excel', 'xlsx', 'spreadsheet', 'sheet', 'table data', "
    "'엑셀', '스프레드시트', '집계표', 'Excel로 받고 싶다' etc. "
    "Each sheet is laid out as: row 1 = title, row 3 = header, row 4+ = data. "
    "Pass an optional `title` per sheet (defaults to sheet name) plus headers and rows. "
    "If an admin uploaded a master template, fill `template_tokens` with values for any "
    "`{{key}}` placeholders the template's cover/summary sheet uses (e.g. "
    "`{'title': '...', 'date': '...', 'author': '...'}`) — the cover sheet is preserved "
    "from the template and only its tokens get substituted. "
    "Charts and formulas are not supported in v1."
)


def _build_dynamic_description() -> str:
    """기본 description 에 현재 템플릿에서 발견된 컨텐츠 토큰 키 안내를 덧붙임.

    빌드 시점에 한 번 — 템플릿이 바뀌면 다음 세션에서 갱신된다 (StructuredTool 은
    factory call 마다 새로 생성). 시스템 자동 채움 키 (date/author/etc.) 는 제외 —
    LLM 이 신경 쓸 키만 노출해 noise 줄임.
    """
    base = _TOOL_DESCRIPTION
    discovered = _discover_template_token_keys()
    content_keys = sorted(discovered - SYSTEM_TOKEN_KEYS)
    if content_keys:
        keys_csv = ", ".join(content_keys)
        base += (
            " IMPORTANT — the admin's Excel template contains these content "
            f"placeholders that you MUST fill via `template_tokens`: {keys_csv}. "
            "Date/author tokens are auto-filled by the system; do not include them."
        )
    return base


def make_create_xlsx(user_id: str) -> StructuredTool:
    """user_id에 바인딩된 create_xlsx 툴을 생성."""

    def _create_xlsx(**kwargs) -> str:
        log.info(
            "create_xlsx called: user_id=%s, template_tokens=%r, sheet_count=%d",
            user_id,
            kwargs.get("template_tokens"),
            len(kwargs.get("sheets") or []),
        )
        content = XlsxContent(**kwargs)
        system_tokens = system_token_values(user_id)
        buf = _build_xlsx(content, system_tokens=system_tokens)
        result = save_to_files(
            user_id=user_id,
            filename=f"{content.filename}.xlsx",
            buffer=buf,
            mime=XLSX_MIME,
        )
        return format_tool_response(result)

    return StructuredTool.from_function(
        func=_create_xlsx,
        name="create_xlsx",
        description=_build_dynamic_description(),
        args_schema=XlsxContent,
    )
