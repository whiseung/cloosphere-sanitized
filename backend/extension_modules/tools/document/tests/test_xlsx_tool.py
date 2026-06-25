"""create_xlsx 툴 테스트."""

import pytest
from extension_modules.tools.document.xlsx_tool import XlsxContent, XlsxSheet
from pydantic import ValidationError


class TestXlsxSchema:
    def test_minimal_valid(self):
        c = XlsxContent(
            filename="data",
            sheets=[XlsxSheet(name="Sheet1", columns=["A", "B"], rows=[["x", "y"]])],
        )
        assert c.sheets[0].name == "Sheet1"

    def test_empty_sheets_rejected(self):
        with pytest.raises(ValidationError):
            XlsxContent(filename="data", sheets=[])

    def test_sheet_name_too_long_rejected(self):
        with pytest.raises(ValidationError):
            XlsxSheet(name="a" * 32, columns=["A"], rows=[])

    def test_row_width_mismatch_rejected(self):
        with pytest.raises(ValidationError, match="컬럼 수 불일치"):
            XlsxSheet(name="S", columns=["A", "B"], rows=[["only_one"]])

    def test_empty_columns_rejected(self):
        with pytest.raises(ValidationError):
            XlsxSheet(name="S", columns=[], rows=[])

    def test_empty_rows_allowed(self):
        s = XlsxSheet(name="S", columns=["A"], rows=[])
        assert s.rows == []

    def test_mixed_cell_types(self):
        s = XlsxSheet(
            name="S",
            columns=["str", "int", "float", "none"],
            rows=[["a", 1, 1.5, None]],
        )
        assert s.rows == [["a", 1, 1.5, None]]


# ── Builder + Factory tests ───────────────────────────────────────────────────

from unittest.mock import MagicMock, patch  # noqa: E402

from extension_modules.tools.document.xlsx_tool import (  # noqa: E402
    XLSX_MIME,
    _build_xlsx,
    make_create_xlsx,
)
from openpyxl import load_workbook  # noqa: E402


class TestXlsxBuilder:
    def test_round_trip_single_sheet(self):
        content = XlsxContent(
            filename="report",
            sheets=[
                XlsxSheet(
                    name="Sales",
                    columns=["Region", "Revenue"],
                    rows=[["KR", 1000], ["JP", 800]],
                )
            ],
        )
        buf = _build_xlsx(content)
        buf.seek(0)
        wb = load_workbook(buf)

        assert wb.sheetnames == ["Sales"]
        ws = wb["Sales"]
        # 레이아웃: row1=제목, row2=공백, row3=헤더, row4~=데이터
        assert ws["A1"].value == "Sales"  # title 미지정 → 시트명 사용
        assert [c.value for c in ws[3]] == ["Region", "Revenue"]
        assert [c.value for c in ws[4]] == ["KR", 1000]
        assert [c.value for c in ws[5]] == ["JP", 800]

    def test_round_trip_multiple_sheets_and_empty_rows(self):
        content = XlsxContent(
            filename="r",
            sheets=[
                XlsxSheet(name="A", columns=["x"], rows=[["1"]]),
                XlsxSheet(name="B", columns=["y", "z"], rows=[]),
            ],
        )
        buf = _build_xlsx(content)
        buf.seek(0)
        wb = load_workbook(buf)
        assert wb.sheetnames == ["A", "B"]
        # 헤더는 row 3
        assert [c.value for c in wb["B"][3]] == ["y", "z"]


class TestMakeCreateXlsx:
    def test_returns_structured_tool_with_correct_name(self):
        tool = make_create_xlsx(user_id="user-1")
        assert tool.name == "create_xlsx"
        # description에 핵심 키워드 포함
        assert "엑셀" in tool.description or "스프레드시트" in tool.description

    def test_tool_invocation_calls_save(self):
        tool = make_create_xlsx(user_id="user-1")
        with patch(
            "extension_modules.tools.document.xlsx_tool.save_to_files"
        ) as mock_save:
            mock_save.return_value = MagicMock(
                file_id="f-1",
                filename="data.xlsx",
                download_url="/api/v1/files/f-1/content",
                size_bytes=512,
            )
            payload = {
                "filename": "data",
                "sheets": [{"name": "S", "columns": ["A"], "rows": [["x"]]}],
            }
            result = tool.invoke(payload)
            mock_save.assert_called_once()
            kwargs = mock_save.call_args.kwargs
            assert kwargs["user_id"] == "user-1"
            assert kwargs["filename"] == "data.xlsx"
            assert kwargs["mime"] == XLSX_MIME


# ── Styling tests ─────────────────────────────────────────────────────────────


class TestXlsxStyling:
    def _build_and_load(self, sheet: XlsxSheet):
        content = XlsxContent(filename="t", sheets=[sheet])
        buf = _build_xlsx(content)
        buf.seek(0)
        return load_workbook(buf)[sheet.name]

    def test_title_row_styled(self):
        # title 미지정 → 시트명, bold + 14pt
        ws = self._build_and_load(
            XlsxSheet(name="요약", columns=["A", "B"], rows=[["x", "y"]])
        )
        assert ws["A1"].value == "요약"
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.size == 14

    def test_explicit_title_used(self):
        ws = self._build_and_load(
            XlsxSheet(name="S", title="2026 매출", columns=["A"], rows=[["x"]])
        )
        assert ws["A1"].value == "2026 매출"

    def test_header_is_bold_and_filled(self):
        # 헤더는 row 3
        ws = self._build_and_load(
            XlsxSheet(name="S", columns=["A", "B"], rows=[["x", "y"]])
        )
        assert ws["A3"].font.bold is True
        assert ws["B3"].font.bold is True
        # 헤더 fill 적용 (옅은 회색)
        assert ws["A3"].fill.fgColor.rgb is not None

    def test_header_has_medium_bottom_border(self):
        ws = self._build_and_load(XlsxSheet(name="S", columns=["A"], rows=[["x"]]))
        # 헤더(row3) 아래쪽 medium, 데이터(row4) 아래쪽 thin
        assert ws["A3"].border.bottom.style == "medium"
        assert ws["A4"].border.bottom.style == "thin"

    def test_data_cells_have_thin_borders_all_sides(self):
        ws = self._build_and_load(XlsxSheet(name="S", columns=["A"], rows=[["x"]]))
        # 데이터는 row 4
        b = ws["A4"].border
        assert b.left.style == "thin"
        assert b.right.style == "thin"
        assert b.top.style == "thin"
        assert b.bottom.style == "thin"

    def test_column_width_autofit_short_content(self):
        ws = self._build_and_load(XlsxSheet(name="S", columns=["A"], rows=[["x"]]))
        # 짧은 내용 → min width 10
        assert ws.column_dimensions["A"].width == 10

    def test_column_width_autofit_long_content(self):
        long_text = "a" * 80
        ws = self._build_and_load(
            XlsxSheet(name="S", columns=["A"], rows=[[long_text]])
        )
        # 80자 → max width 50 으로 캡
        assert ws.column_dimensions["A"].width == 50

    def test_column_width_autofit_korean_doubles(self):
        # 한글 5자 = 시각적 너비 10 → +2 padding = 12 → min 10 보다 큼
        ws = self._build_and_load(
            XlsxSheet(name="S", columns=["서울특별시"], rows=[["부산"]])
        )
        # 헤더 5자 한글 = 시각적 10, +2 padding = 12
        assert ws.column_dimensions["A"].width == 12

    def test_column_width_uses_max_of_header_and_rows(self):
        # 헤더 짧고 데이터 김 → 데이터 기준
        ws = self._build_and_load(
            XlsxSheet(
                name="S",
                columns=["A"],
                rows=[["short"], ["a longer value here"]],
            )
        )
        # 가장 긴 셀 = "a longer value here" = 19자 → +2 = 21
        assert ws.column_dimensions["A"].width == 21
